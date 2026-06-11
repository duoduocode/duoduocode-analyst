"""Export player contribution detectors to Excel (detailed metric rows)."""
import json, sys, os
from typing import Dict, Tuple
sys.path.insert(0, '.')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.collector.api_client import PLAYER_STAT_MAP
from src.engine.player_insights import (
    run_all_detectors,
    D1_METRICS, D2_METRICS, D3_METRICS,
    D9_METRICS, ROLE_METRICS,
)

REVERSE_MAP = {v: k for k, v in PLAYER_STAT_MAP.items()}

# ── English name lookup from type_id ──
# type_id → (developer_name, Chinese name)
TYPE_ID_NAMES: Dict[int, Tuple[str, str]] = {
    # 射门
    42: ("SHOTS_TOTAL", "总射门"),
    86: ("SHOTS_ON_TARGET", "射正"),
    # 进球助攻
    52: ("GOALS", "进球"),
    79: ("ASSISTS", "助攻"),
    580: ("BIG_CHANCES_CREATED", "创造绝佳机会"),
    9706: ("CHANCES_CREATED", "创造机会"),
    # 传球
    80: ("PASSES", "传球总次数"),
    117: ("KEY_PASSES", "关键传球"),
    98: ("TOTAL_CROSSES", "传中总次数"),
    99: ("ACCURATE_CROSSES", "精准传中"),
    1533: ("SUCCESSFUL_CROSSES_PERCENTAGE", "传中成功率"),
    122: ("LONG_BALLS", "长传次数"),
    123: ("LONG_BALLS_WON", "成功长传"),
    27269: ("PASSES_IN_FINAL_THIRD", "进攻三区传球"),
    27270: ("LONG_BALLS_WON_PERCENTAGE", "长传成功率"),
    1584: ("ACCURATE_PASSES_PERCENTAGE", "传球成功率"),
    # 防守
    78: ("TACKLES", "抢断"),
    27267: ("TACKLES_WON", "成功抢断"),
    27268: ("TACKLES_WON_PERCENTAGE", "抢断成功率"),
    100: ("INTERCEPTIONS", "拦截"),
    27271: ("BALL_RECOVERY", "球权回收"),
    97: ("BLOCKED_SHOTS", "封堵射门"),
    110: ("DRIBBLED_PAST", "被过人"),
    # 对抗
    105: ("TOTAL_DUELS", "总对抗"),
    106: ("DUELS_WON", "赢得对抗"),
    1491: ("DUELS_LOST", "输掉对抗"),
    27276: ("DUELS_WON_PERCENTAGE", "对抗成功率"),
    107: ("AERIALS_WON", "赢得空中对抗"),
    27274: ("AERIALS", "空中对抗总次数"),
    27275: ("AERIALS_WON_PERCENTAGE", "空中对抗成功率"),
    # 盘带
    108: ("DRIBBLE_ATTEMPTS", "尝试过人"),
    109: ("SUCCESSFUL_DRIBBLES", "成功过人"),
    # 犯规
    56: ("FOULS", "犯规"),
    96: ("FOULS_DRAWN", "被犯规"),
    115: ("PENALTIES_WON", "赢得点球"),
    # 基础
    120: ("TOUCHES", "触球次数"),
    119: ("MINUTES_PLAYED", "出场时间"),
    # 高阶
    5304: ("EXPECTED_GOALS", "期望进球(xG)"),
    5305: ("EXPECTED_GOALS_ON_TARGET", "射正期望进球(xGOT)"),
    9685: ("SHOOTING_PERFORMANCE", "射门表现(SP)"),
    # 失误
    27273: ("POSSESSION_LOST", "丢失球权"),
}

# Detector metric key → (type_id or None, english_name)
# For computed metrics without a SportMonks type_id, type_id is None
DET_METRIC_MAP: Dict[str, Dict[str, Tuple]] = {}

# D1
for cn, (tid, _) in D1_METRICS.items():
    info = TYPE_ID_NAMES.get(tid, ("", cn))
    DET_METRIC_MAP.setdefault("D1_progression", {})[cn] = (tid, info[0], info[1])

# D2
for cn, (tid, _) in D2_METRICS.items():
    info = TYPE_ID_NAMES.get(tid, ("", cn))
    DET_METRIC_MAP.setdefault("D2_pressing", {})[cn] = (tid, info[0], info[1])

# D3
for cn, (tid, _) in D3_METRICS.items():
    info = TYPE_ID_NAMES.get(tid, ("", cn))
    DET_METRIC_MAP.setdefault("D3_gravity", {})[cn] = (tid, info[0], info[1])

# D4 - custom computed metrics
DET_METRIC_MAP["D4_tempo"] = {
    "传球": (80, "PASSES", "传球总次数"),
    "传球占比": (None, "PASS_SHARE", "传球占比(%)"),
    "准确率": (1584, "ACCURATE_PASSES_PERCENTAGE", "传球成功率"),
    "三区传球": (27269, "PASSES_IN_FINAL_THIRD", "进攻三区传球"),
    "向前比": (None, "FWD_RATIO", "向前传球占比(%)"),
    "回传": (27272, "BACKWARD_PASSES", "回传次数"),
    "触球": (120, "TOUCHES", "触球次数"),
    "触球占比": (None, "TOUCH_SHARE", "触球占比(%)"),
}

# D5 - computed
DET_METRIC_MAP["D5_twoway"] = {
    "进攻排名": (None, "OFF_RANK", "进攻贡献排名"),
    "防守排名": (None, "DEF_RANK", "防守贡献排名"),
    "进攻z": (None, "OFF_ZSCORE", "进攻Z分数"),
    "防守z": (None, "DEF_ZSCORE", "防守Z分数"),
}

# D6 - events
DET_METRIC_MAP["D6_timing"] = {
    "进球": (52, "GOALS", "进球"),
    "助攻": (79, "ASSISTS", "助攻"),
    "点球进球": (111, "PENALTIES_SCORED", "点球进球"),
    "制胜球": (None, "WINNING_GOAL", "制胜球"),
    "绝平球": (None, "EQUALIZER", "绝平球"),
    "首开记录": (None, "FIRST_GOAL", "首开记录"),
    "替补闪电战": (None, "SUPER_SUB", "替补闪电战"),
    "事件加成": (None, "EVENT_BONUSES", "事件加成标签"),
}

# D7 - per-90 computed
DET_METRIC_MAP["D7_efficiency"] = {
    "分钟": (119, "MINUTES_PLAYED", "出场时间"),
    "xG/90": (None, "XG_PER_90", "每90分钟xG"),
    "KP/90": (None, "KP_PER_90", "每90分钟关键传球"),
    "射门/90": (None, "SHOTS_PER_90", "每90分钟射门"),
    "射正/90": (None, "SOT_PER_90", "每90分钟射正"),
    "过人/90": (None, "DRB_PER_90", "每90分钟成功过人"),
    "射门表现/90": (None, "SP_PER_90", "每90分钟射门表现"),
    "xG": (5304, "EXPECTED_GOALS", "期望进球(xG)"),
    "KP": (117, "KEY_PASSES", "关键传球"),
}

# D8
DET_METRIC_MAP["D8_role_deviation"] = {
    "位置": (None, "POSITION", "位置"),
    "偏离分": (None, "DEVIATION", "角色偏离分"),
}

# D9
for cn, (tid, _) in D9_METRICS.items():
    info = TYPE_ID_NAMES.get(tid, ("", cn))
    DET_METRIC_MAP.setdefault("D9_connector", {})[cn] = (tid, info[0], info[1])

# D10
DET_METRIC_MAP["D10_finishing"] = {
    "射门": (42, "SHOTS_TOTAL", "总射门"),
    "xG": (5304, "EXPECTED_GOALS", "期望进球(xG)"),
    "xG/射门": (None, "XG_PER_SHOT", "每射xG"),
    "射正": (86, "SHOTS_ON_TARGET", "射正"),
    "射正率": (None, "SOT_PCT", "射正率(%)"),
}

# D11
DET_METRIC_MAP["D11_xg_deviation"] = {
    "进球": (52, "GOALS", "进球"),
    "xG": (5304, "EXPECTED_GOALS", "期望进球(xG)"),
    "偏差": (None, "XG_DEVIATION", "xG偏差"),
}

# D12
DET_METRIC_MAP["D12_pure_finisher"] = {
    "进球": (52, "GOALS", "进球"),
    "全队进球": (None, "TEAM_GOALS", "全队总进球"),
    "进球占比": (None, "GOAL_SHARE", "进球占比(%)"),
    "xG": (5304, "EXPECTED_GOALS", "期望进球(xG)"),
    "射门": (42, "SHOTS_TOTAL", "总射门"),
    "射正": (86, "SHOTS_ON_TARGET", "射正"),
    "射正率": (None, "SOT_PCT", "射正率(%)"),
}


# ── Detector labels ──
DETECTOR_LABELS = {
    'D1_progression': 'D1 推进价值',
    'D2_pressing': 'D2 防守扫荡',
    'D3_gravity': 'D3 对抗之王',
    'D4_tempo': 'D4 节奏控制/节拍器',
    'D5_twoway': 'D5 双向负荷',
    'D6_timing': 'D6 时机价值',
    'D7_efficiency': 'D7 效率与产量背离',
    'D8_role_deviation': 'D8 角色偏离度',
    'D9_connector': 'D9 连接器',
    'D10_finishing': 'D10 终结质量',
    'D11_xg_deviation': 'D11 xG背离度',
    'D12_pure_finisher': 'D12 纯终结者',
    'D13_prowess': 'D13 终结能力',
}

# ── Detector tags (≤6字, 对优秀球员的评价) ──
DETECTOR_TAGS = {
    'D1_progression': '推进引擎',
    'D2_pressing': '防守铁闸',
    'D3_gravity': '缠斗高手',
    'D4_tempo': '节拍器',
    'D5_twoway': '全能战士',
    'D6_timing': '关键先生',
    'D7_efficiency': '高效输出',
    'D8_role_deviation': '多面手',
    'D9_connector': '串联枢纽',
    'D10_finishing': '射门质量高',
    'D11_xg_deviation': '超预期终结',
    'D12_pure_finisher': '头号火力点',
    'D13_prowess': '终结者',
}

DETECTOR_ORDER = [
    'D1_progression', 'D2_pressing', 'D3_gravity', 'D4_tempo',
    'D5_twoway', 'D6_timing', 'D7_efficiency', 'D8_role_deviation',
    'D9_connector', 'D10_finishing', 'D11_xg_deviation', 'D12_pure_finisher',
    'D13_prowess',
]


def player_to_lineup(p, team_id):
    details = []
    for field_name, value in p.items():
        type_id = REVERSE_MAP.get(field_name)
        if type_id and value is not None:
            details.append({'type_id': type_id, 'data': {'value': value}})
    return {
        'player_id': p['id'], 'player_name': p['name'],
        'team_id': team_id,
        'position_id': p.get('grid', p.get('position_id', 0)),
        'details': details,
        'player': {
            'image_path': p.get('photo_url', ''),
            'position_id': p.get('grid', p.get('position_id', 0)),
        }
    }


def _format_ev(val):
    """Flatten nested evidence dicts to readable values."""
    if isinstance(val, dict) and 'raw' in val:
        raw = val['raw']
        if isinstance(raw, float):
            return f"{raw:.2f}" if raw != int(raw) else str(int(raw))
        return str(raw)
    if isinstance(val, list):
        return ", ".join(str(v) for v in val) if val else "无"
    if isinstance(val, bool):
        return "是" if val else "否"
    if isinstance(val, float):
        return f"{val:.2f}" if val != int(val) else str(int(val))
    return str(val)


def main(match_id=19683241):
    raw_path = f'data/raw/{match_id}/raw_data.json'
    RAW = json.load(open(raw_path, 'r', encoding='utf-8'))

    home_id = RAW['home_team']['id']
    away_id = RAW['away_team']['id']
    hname = RAW['home_team']['name']
    aname = RAW['away_team']['name']

    lineups = [player_to_lineup(p, home_id) for p in RAW['home_players']]
    lineups += [player_to_lineup(p, away_id) for p in RAW['away_players']]

    events = []
    for e in RAW['events']:
        events.append({
            'type_id': e.get('type_id', 0),
            'player_id': e.get('player_id', 0),
            'team_id': e.get('team_id', 0),
            'related_player_id': e.get('related_player_id', 0),
            'minute': e.get('time_elapsed', 0) or e.get('minute', 0),
            'period_id': e.get('period_id', 1),
            'event_type': e.get('event_type', ''),
            'detail': e.get('detail', ''),
            'player_name': e.get('player_name', ''),
            'related_player_name': e.get('related_player_name', ''),
        })

    end_min = 120 if RAW.get('score', {}).get('extratime_home') is not None else 90

    results = run_all_detectors(
        lineups, home_id, away_id,
        RAW['score']['home'], RAW['score']['away'],
        events, end_min,
        home_name=hname, away_name=aname,
    )

    # ── Build rows: one row per player per detector per metric ──
    rows = []  # each: (team, player, detector, scored, metric_raw, metric_zh, value, score, team_rank, overall_rank)

    for attr in DETECTOR_ORDER:
        val = getattr(results, attr)
        label = DETECTOR_LABELS[attr]
        metric_map = DET_METRIC_MAP.get(attr, {})

        if isinstance(val, dict):
            # Per-team dict
            # Build overall ranking
            all_ent = []
            for team, plist in val.items():
                for i, r in enumerate(plist):
                    all_ent.append({'team': team, 'name': r.name, 'score': r.score, 'team_rank': i + 1})
            all_ent.sort(key=lambda x: -x['score'])
            for oi, e in enumerate(all_ent, 1):
                e['overall_rank'] = oi

            for team, plist in val.items():
                team_size = len(plist)
                for i, r in enumerate(plist):
                    overall_r = next(e['overall_rank'] for e in all_ent if e['team'] == team and e['name'] == r.name)
                    tr = i + 1  # team rank numerator only

                    for ev_key, ev_val in r.evidence.items():
                        mapped = metric_map.get(ev_key)
                        if mapped:
                            tid, raw_name, zh_name = mapped
                            if raw_name:
                                metric_label = f"{raw_name} ({zh_name})"
                            else:
                                metric_label = f"{zh_name}"
                        else:
                            metric_label = ev_key

                        value_str = _format_ev(ev_val)
                        rows.append((
                            team,            # 球队
                            r.name,          # 球员
                            label,           # 检测器
                            f"{r.score:.2f}", # 检测器得分
                            metric_label,    # 指标原始名+中文名
                            value_str,       # 指标值
                            tr,              # 全队排名(分子)
                            team_size,       # 全队人数(分母)
                            overall_r,       # 全场排名(分子)
                            len(all_ent),    # 全场人数(分母)
                        ))

                    # If no evidence/metrics, still write a row
                    if not r.evidence:
                        rows.append((
                            team, r.name, label, f"{r.score:.2f}",
                            "-", "-", tr, team_size, overall_r, len(all_ent),
                        ))

        elif isinstance(val, list):
            # D6 timing (cross-team)
            all_ent = [{'name': r.name, 'score': r.score, 'team_rank': i + 1} for i, r in enumerate(val)]
            total = len(all_ent)
            for i, r in enumerate(val):
                for ev_key, ev_val in r.evidence.items():
                    mapped = metric_map.get(ev_key)
                    if mapped:
                        tid, raw_name, zh_name = mapped
                        metric_label = f"{raw_name} ({zh_name})" if raw_name else zh_name
                    else:
                        metric_label = ev_key

                    value_str = _format_ev(ev_val)
                    rows.append((
                        "双方", r.name, label, f"{r.score:.2f}",
                        metric_label, value_str,
                        i + 1, total, i + 1, total,
                    ))

                if not r.evidence:
                    rows.append((
                        "双方", r.name, label, f"{r.score:.2f}",
                        "-", "-", i + 1, total, i + 1, total,
                    ))

    print(f"Total rows: {len(rows)}")

    # ── Create Excel ──
    wb = Workbook()
    ws = wb.active
    ws.title = f"{hname}vs{aname}"

    # Styles
    header_font = Font(name='Microsoft YaHei', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    title_font = Font(name='Microsoft YaHei', bold=True, size=14, color='1E3A5F')
    home_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    away_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    both_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Title
    ws.merge_cells('A1:J1')
    ws['A1'].value = f"球员贡献检测器报告 — {hname} {RAW['score']['home']}-{RAW['score']['away']} {aname}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    headers = ['球队', '球员姓名', '检测器', '检测器得分', '指标', '指标值', '全队排名', '全队总人数', '全场排名', '全场总人数']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Column widths
    for i, w in enumerate([16, 20, 26, 12, 40, 16, 12, 12, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    for row_idx, row_data in enumerate(rows, 4):
        team = row_data[0]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='Microsoft YaHei', size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if team == hname:
                cell.fill = home_fill
            elif team == aname:
                cell.fill = away_fill
            else:
                cell.fill = both_fill

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:J{3 + len(rows)}'

    out_path = f'output/{match_id}_{hname}_vs_{aname}/player_insights_v2.xlsx'
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f'Excel saved: {out_path}')

    # Summary
    from collections import Counter
    det_c = Counter(r[2] for r in rows)
    for d in DETECTOR_ORDER:
        label = DETECTOR_LABELS[d]
        print(f'  {label}: {det_c.get(label, 0)} rows')

    return out_path


if __name__ == '__main__':
    match_id = int(sys.argv[1]) if len(sys.argv) > 1 else 19683241
    main(match_id)
