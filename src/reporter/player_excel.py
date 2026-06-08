"""
Player Contribution Excel Exporter v6

Generates .xlsx files with structured sheets per match + cross-match role summary.
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers,
)
from openpyxl.utils import get_column_letter

from src.engine.player_insights_v6 import (
    PlayerInsightV6, ContributionScore,
    C1_DISPLAY_ORDER, C1_DISPLAY_NAMES,
    C2_DISPLAY_ORDER, C2_DISPLAY_NAMES,
    C3_DISPLAY_ORDER, C3_DISPLAY_NAMES,
    C4_DISPLAY_ORDER, C4_DISPLAY_NAMES,
    C5_DISPLAY_ORDER, C5_DISPLAY_NAMES,
    C_SUBTITLES, C_DIM_NAMES,
)

# ── Styles ──
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=13, color="1F4E79")
SUB_FONT = Font(name="Microsoft YaHei", bold=True, size=10, color="1F4E79")
NORMAL_FONT = Font(name="Microsoft YaHei", size=9)
BOLD_FONT = Font(name="Microsoft YaHei", bold=True, size=9)
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RED_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GK_FILL = PatternFill(start_color="E2D9F3", end_color="E2D9F3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _style_header(ws, row: int, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_data_cell(cell, value, bold=False, fill=None):
    """Apply formatting to a data cell."""
    cell.value = value
    cell.font = BOLD_FONT if bold else NORMAL_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill


def _safe_float(v, default=""):
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return round(float(v), 2) if abs(float(v)) >= 0.01 else round(float(v), 3)
    return default


def _dim_highlight(zscore: float, rank: int) -> Optional[PatternFill]:
    """Green for top-3, red for negative."""
    if zscore > 1.5 and rank <= 3:
        return GREEN_FILL
    if zscore < -1.0:
        return RED_FILL
    if zscore > 0.5:
        return YELLOW_FILL
    return None


# ═══════════════════════════════════════════════════════════════
# Sheet 1: Summary
# ═══════════════════════════════════════════════════════════════

def _write_summary_sheet(ws, insights: list[PlayerInsightV6], match_name: str, score: str):
    """Main player contribution summary sheet."""
    ws.title = "球员贡献概要"

    # Title
    ws.merge_cells("A1:Q1")
    ws["A1"] = f"球员贡献分析 — {match_name}  |  比分 {score}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # Headers
    headers = ["位", "#", "球员", "分钟", "C1进攻", "C2推进", "C3控制", "C4防守", "C5对抗", "C6事件",
               "进球", "xG", "助攻", "传球", "抢断", "关键传", "角色"]
    ncols = len(headers)
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header(ws, 3, ncols)

    # Subtitle row
    subtitles = ["", "", "", "",
                 C_SUBTITLES.get("C1", ""), C_SUBTITLES.get("C2", ""),
                 C_SUBTITLES.get("C3", ""), C_SUBTITLES.get("C4", ""),
                 C_SUBTITLES.get("C5", ""), "", "", "", "", "", "", "", ""]
    for c, st in enumerate(subtitles, 1):
        cell = ws.cell(row=4, column=c, value=st)
        cell.font = Font(name="Microsoft YaHei", italic=True, size=8, color="666666")
        cell.alignment = CENTER

    def _sort_key(p):
        dims = [p.contributions.get(k, ContributionScore(0, 99, 0, "", {})).zscore
                for k in ("C1", "C2", "C3", "C4", "C5")]
        return -max(dims, default=0)

    row = 5
    for pi in sorted(insights, key=_sort_key):
        c1 = pi.contributions.get("C1", ContributionScore(0, 99, 0, "", {}))
        c2 = pi.contributions.get("C2", ContributionScore(0, 99, 0, "", {}))
        c3 = pi.contributions.get("C3", ContributionScore(0, 99, 0, "", {}))
        c4 = pi.contributions.get("C4", ContributionScore(0, 99, 0, "", {}))
        c5 = pi.contributions.get("C5", ContributionScore(0, 99, 0, "", {}))
        c6 = pi.contributions.get("C6", ContributionScore(0, 0, 0, "", {}))

        is_gk = pi.pos == "G"
        vals = [
            pi.pos if not is_gk else "GK",
            pi.number,
            pi.name,
            pi.minutes,
            _safe_float(c1.zscore) if not is_gk else "GK",
            _safe_float(c2.zscore) if not is_gk else "GK",
            _safe_float(c3.zscore),
            _safe_float(c4.zscore) if not is_gk else "GK",
            _safe_float(c5.zscore) if not is_gk else "GK",
            _safe_float(c6.zscore) if c6.zscore != 0 else "-",
            pi.sv("goals", 0) if not is_gk else "-",
            _safe_float(pi.sv("xg", 0)) if not is_gk else "-",
            pi.sv("assists", 0) if not is_gk else "-",
            int(pi.sv("passes_total", 0)),
            int(pi.sv("tackles_total", 0)),
            int(pi.sv("passes_key", 0)),
            pi.role.name if pi.role else "-",
        ]

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = THIN_BORDER
            cell.font = NORMAL_FONT
            cell.alignment = CENTER

            # Per-column highlights
            if c in (5, 6, 7, 8, 9) and isinstance(v, (int, float)):
                fill = _dim_highlight(float(v), c1.rank if c == 5 else c2.rank if c == 6 else c3.rank if c == 7 else c4.rank if c == 8 else c5.rank)
                if fill:
                    cell.fill = fill
            if is_gk:
                cell.fill = GK_FILL

        # Bold name
        ws.cell(row=row, column=3).font = BOLD_FONT
        ws.cell(row=row, column=3).alignment = LEFT

        row += 1

    # Column widths
    widths = [4, 4, 16, 6, 7, 7, 7, 7, 7, 7, 5, 6, 5, 6, 5, 6, 10]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.auto_filter.ref = f"A3:Q{row - 1}"
    ws.freeze_panes = "A5"


# ═══════════════════════════════════════════════════════════════
# Sheet 2-6: Dimension Detail
# ═══════════════════════════════════════════════════════════════

def _write_dim_sheet(ws, insights: list[PlayerInsightV6],
                      dim_key: str, dim_label: str,
                      display_order: list[str], display_names: dict[str, str]):
    """Write one dimension detail sheet."""
    ws.title = dim_label

    outfield = [pi for pi in insights if pi.pos != "G"]
    if not outfield:
        return

    headers = ["排名", "球员", "Z综合"] + [display_names.get(k, k) for k in display_order]
    ncols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    dim_full = C_DIM_NAMES.get(dim_key, dim_label)
    sub = C_SUBTITLES.get(dim_key, "")
    ws["A1"] = f"{dim_full} — {sub}" if sub else dim_full
    ws["A1"].font = SUB_FONT

    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header(ws, 3, ncols)

    sorted_insights = sorted(outfield, key=lambda pi: -pi.contributions.get(dim_key, ContributionScore(0, 99, 0, "", {})).zscore)

    row = 4
    for rank, pi in enumerate(sorted_insights, 1):
        c = pi.contributions.get(dim_key)
        if c is None:
            continue

        vals = [rank, pi.name, _safe_float(c.zscore)]
        for k in display_order:
            raw = pi.raw_stats.get(k, 0)
            if raw is None:
                raw = 0
            # Get metric ranks
            mdata = c.raw_metrics.get(k, {})
            tr = mdata.get("_team_rank", "")
            mr = mdata.get("_match_rank", "")
            rank_suffix = f" (队{tr}/场{mr})" if tr and mr else ""

            if isinstance(raw, float):
                val_str = f"{round(raw, 3) if abs(raw) < 10 else round(raw, 1)}{rank_suffix}"
            elif isinstance(raw, int):
                val_str = f"{raw}{rank_suffix}"
            else:
                val_str = str(raw)
            vals.append(val_str)

        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = THIN_BORDER
            cell.font = NORMAL_FONT
            cell.alignment = CENTER
            if col == 2:
                cell.font = BOLD_FONT
                cell.alignment = LEFT
            if col == 3 and isinstance(v, (int, float)):
                fill = _dim_highlight(float(v), rank)
                if fill:
                    cell.fill = fill

        row += 1

    # Column widths (wider for rank info)
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16 if c > 2 else 8

    ws.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════
# Sheet 7: C6 Key Events
# ═══════════════════════════════════════════════════════════════

def _write_c6_sheet(ws, insights: list[PlayerInsightV6]):
    ws.title = "C6关键事件"

    headers = ["球员", "总分", "制胜球", "绝杀", "绝平球", "首开记录",
               "点球进球", "制造点球", "超级替补",
               "PK战进球", "PK战射失", "标签"]
    ncols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = "C6 关键事件明细"
    ws["A1"].font = SUB_FONT

    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header(ws, 3, ncols)

    sorted_pis = sorted(insights, key=lambda pi: -pi.contributions.get("C6", ContributionScore(0, 0, 0, "", {})).zscore)

    row = 4
    for pi in sorted_pis:
        eb = pi.event_bonus
        if eb is None:
            continue
        c6 = pi.contributions.get("C6")
        if c6 is None or c6.zscore == 0:
            continue

        vals = [
            pi.name,
            c6.zscore,
            "Y" if eb.winning_goal else "-",
            "Y" if eb.late_winner else "-",
            "Y" if eb.equalizer else "-",
            "Y" if eb.first_goal else "-",
            "Y" if eb.scored_penalty else "-",
            "Y" if eb.won_penalty else "-",
            "Y" if eb.super_sub else "-",
            "Y" if eb.pen_shootout_goal else "-",
            "Y" if eb.pen_shootout_miss else "-",
            eb.c6_label(),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = THIN_BORDER
            cell.font = NORMAL_FONT
            cell.alignment = CENTER
            if c == 1:
                cell.font = BOLD_FONT
                cell.alignment = LEFT

        row += 1

    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11 if c > 1 else 14


# ═══════════════════════════════════════════════════════════════
# Sheet 8: C7 Goalkeeper
# ═══════════════════════════════════════════════════════════════

def _write_c7_sheet(ws, insights: list[PlayerInsightV6]):
    gks = [pi for pi in insights if pi.pos == "G"]
    if not gks:
        return

    ws.title = "C7门将"

    headers = ["球员", "球队", "综合分", "扑救", "禁区扑救", "失球", "摘高球", "击球",
               "传球%", "长传成功", "长传%", "丢球权", "致命失误", "角色"]
    ncols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = "C7 门将贡献"
    ws["A1"].font = SUB_FONT

    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header(ws, 3, ncols)

    row = 4
    for pi in sorted(gks, key=lambda p: -p.contributions.get("C7", ContributionScore(0, 0, 0, "", {})).zscore):
        c7 = pi.contributions.get("C7")
        if c7 is None:
            continue
        raw = c7.raw_metrics
        vals = [
            pi.name,
            pi.team_name,
            _safe_float(c7.zscore),
            raw.get("saves", 0),
            raw.get("saves_inside_box", 0),
            raw.get("goalkeeper_goals_conceded", 0),
            raw.get("good_high_claim", 0),
            raw.get("punches", 0),
            raw.get("passes_accuracy", 0),
            raw.get("long_balls_won", 0),
            raw.get("long_balls_won_pct", 0),
            raw.get("possession_lost", 0),
            raw.get("error_lead_to_goal", 0),
            c7.label,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = THIN_BORDER
            cell.font = NORMAL_FONT
            cell.alignment = CENTER
            if c == 1:
                cell.font = BOLD_FONT
                cell.alignment = LEFT

        row += 1

    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11


# ═══════════════════════════════════════════════════════════════
# Sheet 9: Role Narratives
# ═══════════════════════════════════════════════════════════════

def _write_role_sheet(ws, insights: list[PlayerInsightV6]):
    ws.title = "角色叙事"

    headers = ["球员", "位置", "球队", "角色", "置信度", "LLM分析 / 角色叙事"]
    ncols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = "第二层：球员场上作用分析"
    ws["A1"].font = SUB_FONT

    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header(ws, 3, ncols)

    row = 4
    for pi in insights:
        if pi.role or pi.llm_summary:
            narrative = pi.llm_summary or (pi.role.narrative if pi.role else "")
            role_name = pi.role.name if pi.role else "-"
            confidence = _safe_float(pi.role.confidence) if pi.role else "-"

            vals = [
                pi.name, pi.pos, pi.team_name,
                role_name, confidence, narrative,
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.border = THIN_BORDER
                cell.font = NORMAL_FONT
                cell.alignment = LEFT_TOP if c == 6 else CENTER
                if pi.llm_summary and c == 6:
                    cell.fill = YELLOW_FILL
            row += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 80


# ═══════════════════════════════════════════════════════════════
# Public API
# ═══════════════════════════════════════════════════════════════

def export_match_excel(
    insights: list[PlayerInsightV6],
    match_id: str,
    match_name: str,
    score: str,
    output_path: str,
):
    """Export full player contribution analysis for one match to Excel."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Summary
    ws1 = wb.create_sheet("概要")
    _write_summary_sheet(ws1, insights, match_name, score)

    # Dimension details
    _write_dim_sheet(wb.create_sheet("C1进攻"), insights, "C1", "C1 进攻贡献",
                     C1_DISPLAY_ORDER, C1_DISPLAY_NAMES)
    _write_dim_sheet(wb.create_sheet("C2推进"), insights, "C2", "C2 推进贡献",
                     C2_DISPLAY_ORDER, C2_DISPLAY_NAMES)
    _write_dim_sheet(wb.create_sheet("C3控制"), insights, "C3", "C3 控制贡献",
                     C3_DISPLAY_ORDER, C3_DISPLAY_NAMES)
    _write_dim_sheet(wb.create_sheet("C4防守"), insights, "C4", "C4 防守贡献",
                     C4_DISPLAY_ORDER, C4_DISPLAY_NAMES)
    _write_dim_sheet(wb.create_sheet("C5对抗"), insights, "C5", "C5 对抗贡献",
                     C5_DISPLAY_ORDER, C5_DISPLAY_NAMES)

    # Events & GK
    _write_c6_sheet(wb.create_sheet("C6事件"), insights)
    _write_c7_sheet(wb.create_sheet("C7门将"), insights)
    _write_role_sheet(wb.create_sheet("角色叙事"), insights)

    wb.save(output_path)


def export_cross_match_role_summary(
    all_results: dict[str, list[PlayerInsightV6]],
    output_path: str,
):
    """Export cross-match role distribution summary."""
    wb = Workbook()
    ws = wb.active
    ws.title = "角色汇总"

    ws["A1"] = "球员角色分布 — 三场汇总"
    ws["A1"].font = TITLE_FONT

    # Collect all roles
    all_roles: dict[str, dict[str, list[str]]] = {}
    match_ids = sorted(all_results.keys())

    for mid in match_ids:
        insights = all_results[mid]
        for pi in insights:
            if pi.role and pi.pos != "G":
                role = pi.role.name
                if role not in all_roles:
                    all_roles[role] = {}
                if mid not in all_roles[role]:
                    all_roles[role][mid] = []
                all_roles[role][mid].append(f"{pi.name} ({pi.team_name}, {pi.minutes}')")

    role_order = [
        "蹲坑中卫", "出球后卫", "进攻型边卫", "全能后卫",
        "节拍器", "扫荡后腰", "全能中场", "进攻组织者", "串联枢纽",
        "终结者", "全能中锋", "边路突击手", "伪九号",
    ]
    sorted_roles = [r for r in role_order if r in all_roles]
    sorted_roles += [r for r in sorted(all_roles) if r not in role_order]

    # Headers
    headers = ["角色"] + [str(m) for m in match_ids]
    npats = len(match_ids) + 1
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header(ws, 3, npats)

    row = 4
    for role in sorted_roles:
        ws.cell(row=row, column=1, value=role).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
        for ci, mid in enumerate(match_ids, 2):
            players = all_roles[role].get(mid, [])
            val = ", ".join(players) if players else "—"
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = THIN_BORDER
            cell.font = NORMAL_FONT
            cell.alignment = LEFT_TOP
        row += 1

    ws.column_dimensions["A"].width = 14
    for ci in range(2, npats + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 55

    wb.save(output_path)
