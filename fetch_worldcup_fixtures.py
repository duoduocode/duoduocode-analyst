"""
获取 2026 世界杯小组赛所有比赛 fixture_id

用法:
    python fetch_worldcup_fixtures.py

输出:
    output/2026worldcup/worldcup_group_stage.xlsx
    四列: 小组, 比赛双方, 比赛时间, fixture_id
"""

import argparse
import logging
import os
import sys
from datetime import datetime, timedelta

import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

GROUP_STAGE_START = "2026-06-11"
GROUP_STAGE_END = "2026-06-28"

# SportMonks 中 2026 世界杯的正确 league_id
WORLD_CUP_LEAGUE_ID = 732


def load_config(path: str = "config.yaml") -> dict:
    """与 run.py 一致的配置加载方式，替换 ${ENV_VAR} 占位符"""
    with open(path, "r", encoding="utf-8") as f:
        raw = f.read()
    for key, value in os.environ.items():
        placeholder = "${" + key + "}"
        raw = raw.replace(placeholder, value)
    return yaml.safe_load(raw)


def is_group_stage(stage_name: str) -> bool:
    return "group" in (stage_name or "").lower()


def main():
    parser = argparse.ArgumentParser(description="获取世界杯小组赛 fixtures")
    parser.add_argument("--config", default="config.yaml", help="配置文件路径")
    parser.add_argument("--start-date", default=GROUP_STAGE_START)
    parser.add_argument("--end-date", default=GROUP_STAGE_END)
    parser.add_argument("--league-id", type=int, default=WORLD_CUP_LEAGUE_ID)
    args = parser.parse_args()

    config = load_config(args.config)
    league_id = args.league_id

    # 生成日期列表
    start_dt = datetime.strptime(args.start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(args.end_date, "%Y-%m-%d")
    dates = [(start_dt + timedelta(days=i)).strftime("%Y-%m-%d")
             for i in range((end_dt - start_dt).days + 1)]

    logger.info(f"日期范围: {args.start_date} → {args.end_date} ({len(dates)} 天)")
    logger.info(f"联赛 ID: {league_id}")

    api_token = config["sportmonks"]["api_token"]
    base_url = config["sportmonks"].get("base_url", "https://api.sportmonks.com/v3/football")
    session = requests.Session()
    session.trust_env = False

    all_fixtures = []

    for date_str in dates:
        try:
            params = {
                "api_token": api_token,
                "include": "participants;group;stage",
                "filters": f"fixtureLeagues:{league_id}",
            }
            url = f"{base_url}/fixtures/date/{date_str}"
            resp = session.get(url, params=params, timeout=60)
            resp.raise_for_status()
            data = resp.json()
            if data.get("error"):
                logger.warning(f"  {date_str}: API error - {data['error']}")
                continue
            fixtures = data.get("data", [])
            if isinstance(fixtures, dict):
                fixtures = [fixtures]
            logger.info(f"  {date_str}: {len(fixtures)} 场比赛")
            all_fixtures.extend(fixtures)
        except Exception as e:
            logger.warning(f"  {date_str}: 获取失败 - {e}")

    logger.info(f"共获取 {len(all_fixtures)} 场比赛 (未过滤)")

    # 过滤小组赛
    group_fixtures = []
    for f in all_fixtures:
        stage = f.get("stage", {}) or {}
        stage_name = stage.get("name", "")
        group = f.get("group", {}) or {}
        if is_group_stage(stage_name) or bool(group.get("id")):
            group_fixtures.append(f)

    logger.info(f"过滤后小组赛: {len(group_fixtures)} 场")

    if not group_fixtures:
        logger.warning("未找到任何小组赛数据。")
        return 1

    # 构建输出数据
    rows = []
    for f in group_fixtures:
        fixture_id = f.get("id", "")

        group = f.get("group", {}) or {}
        group_name = group.get("name", "")

        participants = f.get("participants", []) or []
        if len(participants) >= 2:
            home_name = participants[0].get("name", "?")
            away_name = participants[1].get("name", "?")
            teams = f"{home_name} vs {away_name}"
        else:
            teams = "?? vs ??"

        starting_at = f.get("starting_at", "")
        if starting_at:
            try:
                dt = datetime.fromisoformat(starting_at.replace("Z", "+00:00"))
                match_time = dt.strftime("%Y-%m-%d %H:%M")
            except (ValueError, TypeError):
                match_time = str(starting_at)
        else:
            match_time = ""

        rows.append({
            "group": group_name or "?",
            "teams": teams,
            "time": match_time,
            "fixture_id": fixture_id,
        })

    # 按小组和时间排序
    rows.sort(key=lambda r: (r["group"], r["time"]))

    # 输出到 Excel
    output_dir = "output/2026worldcup"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "worldcup_group_stage.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "世界杯小组赛"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["小组", "比赛双方", "比赛时间", "fixture_id"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(["group", "teams", "time", "fixture_id"], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col_idx in (1, 3, 4) else "left")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.freeze_panes = "A2"

    wb.save(output_path)
    logger.info(f"已保存: {output_path} ({len(rows)} 场小组赛)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
