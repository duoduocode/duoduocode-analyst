"""
Batch compute Finishing Prowess (终结能力) for all 2025/26 UCL knockout fixtures.

Usage:
    python batch_finishing_prowess.py                    # full run (fetch + compute)
    python batch_finishing_prowess.py --skip-fetch       # use cached raw_data.json
    python batch_finishing_prowess.py --fixture 19683241 --skip-fetch  # single fixture test
"""

from __future__ import annotations

import os, sys, json, argparse
from pathlib import Path

for k in ['HTTP_PROXY', 'HTTPS_PROXY', 'http_proxy', 'https_proxy', 'NO_PROXY']:
    os.environ.pop(k, None)

import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, '.')
from src.collector.api_client import (
    SportMonksClient, fetch_all, PLAYER_STAT_MAP,
)
from src.engine.player_insights import (
    PlayerData, EventBonuses,
    detect_finishing_prowess,
    compute_event_bonuses, classify_position,
    run_all_detectors,
)

TOKEN = "w6PU0YE99DrKPnOY43xonH5Bf7S279Ysvzzj6LZaAqnOKusDEIAUTbaNUzDr"
SEASON_ID = 25580
REVERSE_MAP = {v: k for k, v in PLAYER_STAT_MAP.items()}

KO_STAGES = [
    "Play-offs", "Knockout Round Play-offs", "8th Finals",
    "Quarter-finals", "Semi-finals", "Final",
]


# ── Fetch ──
def get_ko_fixture_ids() -> list[dict]:
    s = requests.Session()
    s.trust_env = False
    proxies = {'http': None, 'https': None}

    r = s.get(f'https://api.sportmonks.com/v3/football/stages/seasons/{SEASON_ID}',
              params={'api_token': TOKEN}, proxies=proxies)
    stages = r.json().get('data', [])

    fixtures = []
    for st in stages:
        sname = st.get('name', '')
        if sname not in KO_STAGES:
            continue
        r2 = s.get(f'https://api.sportmonks.com/v3/football/stages/{st["id"]}',
                  params={'api_token': TOKEN, 'include': 'fixtures.participants'}, proxies=proxies)
        for f in r2.json().get('data', {}).get('fixtures', []):
            parts = f.get('participants', [])
            h = parts[0].get('name', '?') if len(parts) > 0 else '?'
            a = parts[1].get('name', '?') if len(parts) > 1 else '?'
            fixtures.append({'id': f['id'], 'stage': sname, 'home': h, 'away': a})

    print(f"Found {len(fixtures)} KO fixtures across {len(KO_STAGES)} stages")
    for f in fixtures:
        print(f"  [{f['stage']}] F{f['id']}: {f['home']} vs {f['away']}")
    return fixtures


# ── Convert raw_data → lineups (reuse export_player_insights.py logic) ──
def player_to_lineup(p: dict, team_id: int) -> dict:
    details = []
    for field_name, value in p.items():
        type_id = REVERSE_MAP.get(field_name)
        if type_id and value is not None:
            details.append({'type_id': type_id, 'data': {'value': value}})
    return {
        'player_id': p.get('id', 0), 'player_name': p.get('name', ''),
        'team_id': team_id,
        'position_id': p.get('grid', p.get('position_id', 0)),
        'details': details,
        'player': {
            'image_path': p.get('photo_url', ''),
            'position_id': p.get('grid', p.get('position_id', 0)),
        }
    }


def raw_to_events(raw: dict) -> list[dict]:
    return [{
        'type_id': e.get('type_id', 0),
        'minute': e.get('time_elapsed', 0) or e.get('minute', 0),
        'period_id': e.get('period_id', 1),
        'team_id': e.get('team_id', e.get('participant_id', 0)),
        'player_name': e.get('player_name', ''),
        'related_player_name': e.get('related_player_name', ''),
        'event_type': e.get('event_type', ''),
        'detail': e.get('detail', ''),
        'participant_id': e.get('team_id', e.get('participant_id', 0)),
    } for e in raw.get('events', [])]


# ── Excel ──
HEADER_FILL = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
HEADER_FONT = Font(name='Microsoft YaHei', bold=True, size=10, color='FFFFFF')
TITLE_FONT = Font(name='Microsoft YaHei', bold=True, size=14, color='1E3A5F')
DATA_FONT = Font(name='Microsoft YaHei', size=10)
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
HOME_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
AWAY_FILL = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
GOLD = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
SILVER = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
BRONZE = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')


def _fm(v):
    if v is None: return "-"
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else (f"{v:.4f}" if abs(v) < 0.01 else f"{v:.2f}")
    return str(v)


def output_excel(team_results: dict, fixture: dict, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "终结能力"
    hname, aname = fixture['home'], fixture['away']

    ws.merge_cells('A1:O1')
    ws['A1'].value = f"终结能力 — {hname} {fixture['score']} {aname} | {fixture['stage']} | UCL 2025/26"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['球队', '球员', '终结分', '进球', 'xG', '超预期', 'xG/射门',
               '射门', '射正', '射正率%', '转化率%', '进球占比%', '中框', '关键标签', '射门表现']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER

    for i, w in enumerate([16, 22, 9, 7, 8, 8, 9, 7, 7, 8, 8, 10, 7, 24, 9], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    all_ent = []
    for tn, res in team_results.items():
        for r in res:
            all_ent.append((tn, r))
    all_ent.sort(key=lambda x: -x[1].score)

    for rank, (team, r) in enumerate(all_ent, 1):
        e = r.evidence
        vals = [team, r.name,
                _fm(e.get('终结分', r.score)), _fm(e.get('进球')),
                _fm(e.get('xG')), _fm(e.get('超预期')),
                _fm(e.get('xG/射门')), _fm(e.get('射门')),
                _fm(e.get('射正')), _fm(e.get('射正率')),
                _fm(e.get('转化率')), _fm(e.get('进球占比')),
                _fm(e.get('中框')), e.get('关键标签', '-'),
                _fm(e.get('射门表现'))]
        fill = HOME_FILL if team == hname else AWAY_FILL
        if rank == 1: fill = GOLD
        elif rank == 2: fill = SILVER
        elif rank == 3: fill = BRONZE
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=rank + 3, column=c, value=v)
            cell.font = DATA_FONT; cell.fill = fill; cell.border = BORDER
            cell.alignment = Alignment(horizontal='center' if c > 2 else 'left')

    ws.freeze_panes = 'C4'
    ws.auto_filter.ref = f"A3:O{len(all_ent) + 3}"
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  -> {out_path}")


# ── Process single fixture ──
def process_fixture(fixture: dict, config: dict, skip_fetch: bool = False) -> bool:
    fid = fixture['id']
    hname_p = fixture['home']
    aname_p = fixture['away']
    stage = fixture['stage']
    raw_path = f'data/raw/{fid}/raw_data.json'

    # Fetch / load
    if not skip_fetch or not os.path.exists(raw_path):
        print(f"[{fid}] Fetching: {hname_p} vs {aname_p}")
        try:
            raw = fetch_all(fid, config['sportmonks'])
            Path(raw_path).parent.mkdir(parents=True, exist_ok=True)
            serialized = {}
            for k in ['match_id','fixture_id','home_team','away_team','home_coach','away_coach',
                       'score','period_scores','status','home_stats','away_stats',
                       'home_players','away_players','events','periods','home_lineup',
                       'away_lineup','trends','timeline','formations','stage_info','venue_info']:
                val = getattr(raw, k, None)
                if val is not None:
                    if hasattr(val, '__dict__'):
                        val = val.__dict__
                    elif isinstance(val, list):
                        val = [{**v.__dict__} if hasattr(v, '__dict__') else v for v in val]
                    serialized[k] = val
            Path(raw_path).write_text(json.dumps(serialized, ensure_ascii=False, indent=2, default=str),
                                     encoding='utf-8')
        except Exception as e:
            print(f"[{fid}] Fetch FAILED: {e}")
            return False
    else:
        print(f"[{fid}] Cached: {hname_p} vs {aname_p}")

    # Load
    raw = json.loads(Path(raw_path).read_text(encoding='utf-8'))
    home_id = raw['home_team']['id']
    away_id = raw['away_team']['id']
    hname = raw['home_team'].get('name', '') or hname_p
    aname = raw['away_team'].get('name', '') or aname_p
    score_h = raw.get('score', {}).get('home', 0)
    score_a = raw.get('score', {}).get('away', 0)
    has_et = raw.get('score', {}).get('extratime_home') is not None
    end_min = 120 if has_et else 90

    out_dir = f"output/{fid}_{hname.replace(' ', '_')}_vs_{aname.replace(' ', '_')}"
    out_file = f'{out_dir}/finishing_prowess.xlsx'

    # Convert to lineups
    lineups = [player_to_lineup(p, home_id) for p in raw.get('home_players', [])]
    lineups += [player_to_lineup(p, away_id) for p in raw.get('away_players', [])]
    events = raw_to_events(raw)

    # Build PlayerData per team
    pos_cache = {}
    def build_players(plist: list[dict], tid: int, tname: str, filter_team: bool = True) -> list[PlayerData]:
        result = []
        for p in plist:
            if filter_team and p.get('team_id') != tid:
                continue
            pid = p.get('player_id', 0)
            if not pid: continue
            pname = p.get('player_name', f"Player #{pid}")
            if pid not in pos_cache:
                pos_id = p.get('position_id', 0)
                photo = p.get('player', {}).get('image_path', '')
                pos_cache[pid] = (classify_position(pos_id), pos_id, photo)
            pos, pos_id, photo = pos_cache[pid]
            stats = {det['type_id']: det['data'].get('value') for det in p.get('details', [])}
            result.append(PlayerData(player_id=pid, name=pname, position_id=pos_id,
                                     pos=pos, team_name=tname, stats=stats, photo_url=photo))
        result.sort(key=lambda x: -(x.sv(119) or 0))
        return result

    home_players = build_players(lineups, home_id, hname)
    away_players = build_players(lineups, away_id, aname)

    # Event bonuses
    bonuses = compute_event_bonuses(events, home_id, away_id, score_h, score_a, end_min)

    # Run detectors
    home_res = detect_finishing_prowess(home_players, score_h, bonuses)
    away_res = detect_finishing_prowess(away_players, score_a, bonuses)

    # Output
    fixture_info = {'id': fid, 'home': hname, 'away': aname,
                    'score': f"{score_h}-{score_a}", 'stage': stage}
    output_excel({hname: home_res, aname: away_res}, fixture_info, out_file)
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--skip-fetch', action='store_true')
    parser.add_argument('--fixture', type=int, default=0)
    args = parser.parse_args()

    with open('config.yaml', 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
    config['sportmonks']['api_token'] = TOKEN

    os.environ['SPORTMONKS_API_TOKEN'] = TOKEN
    for k in ['HTTP_PROXY', 'HTTPS_PROXY', 'http_proxy', 'https_proxy']:
        os.environ.pop(k, None)

    if args.fixture:
        fixtures = [{'id': args.fixture, 'stage': 'Final', 'home': '?', 'away': '?'}]
    else:
        fixtures = get_ko_fixture_ids()

    ok = 0
    for i, f in enumerate(fixtures, 1):
        print(f"\n[{i}/{len(fixtures)}] F{f['id']}")
        if process_fixture(f, config, skip_fetch=args.skip_fetch):
            ok += 1
    print(f"\nDone: {ok}/{len(fixtures)}")


if __name__ == '__main__':
    main()
