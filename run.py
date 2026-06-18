import argparse
import json
import logging
import os
from pathlib import Path

import yaml

from src.collector.api_client import SportMonksClient, fetch_all
from src.engine.metrics import compute_all

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)


def load_config(path: str = "config.yaml") -> dict:
    with open(path, "r", encoding="utf-8") as f:
        raw = f.read()
    for key, value in os.environ.items():
        placeholder = "${" + key + "}"
        raw = raw.replace(placeholder, value)
    return yaml.safe_load(raw)


def generate_narrative(
    raw, computed, signals, trend_analysis, llm_config: dict
) -> str:
    from src.composer.data_builder import build_narrative
    from src.composer.prompt_loader import PromptLoader
    from src.generator.llm_client import LLMClient

    llm = LLMClient(llm_config)
    pl = PromptLoader("prompts")

    logger.info("Building narrative prompt...")
    sys_p, user_p = build_narrative(raw, computed, signals, trend_analysis, pl)

    logger.info("Calling LLM for narrative...")
    text = llm.generate(sys_p, user_p)
    return text


def generate_all_visuals(raw, computed, visual_config: dict, output_dir: str, sub_impacts=None) -> dict:
    from src.engine.metrics import _stat
    from src.visualizer.momentum import plot_momentum_curve_v3
    from src.visualizer.pass_network import plot_pass_network
    from src.visualizer.efficiency import plot_efficiency_comparison
    from src.visualizer.subs import plot_sub_impacts_v3
    from src.visualizer.xg_hist import plot_xg_histogram

    dpi = visual_config.get("dpi", 150)
    images_dir = Path(output_dir) / "images"
    images_dir.mkdir(parents=True, exist_ok=True)

    result = {}

    hs = raw.home_stats
    aws = raw.away_stats
    home_xg = sum(p.xg for p in raw.home_players) if raw.home_players else float(_stat(hs, "Expected Goals", default=0))
    away_xg = sum(p.xg for p in raw.away_players) if raw.away_players else float(_stat(aws, "Expected Goals", default=0))

    # Efficiency comparison chart (replaces shot map)
    result["efficiency"] = plot_efficiency_comparison(
        raw, str(images_dir / "01_efficiency.png"), dpi=dpi,
    )

    # Momentum trends v3 (uses actual trends data)
    result["momentum"] = plot_momentum_curve_v3(
        raw, str(images_dir / "02_momentum.png"), dpi=dpi,
    )

    if raw.home_lineup:
        result["pass_home"] = plot_pass_network(
            raw.home_lineup.players,
            raw.home_players,
            raw.home_lineup.formation,
            raw.home_team.name,
            str(images_dir / "03a_pass_home.png"),
            dpi=dpi,
        )
    else:
        result["pass_home"] = ""

    if raw.away_lineup:
        result["pass_away"] = plot_pass_network(
            raw.away_lineup.players,
            raw.away_players,
            raw.away_lineup.formation,
            raw.away_team.name,
            str(images_dir / "03b_pass_away.png"),
            dpi=dpi,
        )
    else:
        result["pass_away"] = ""

    # Sub impact multi-metric chart (replaces old subs chart)
    if sub_impacts:
        result["subs"] = plot_sub_impacts_v3(
            raw, sub_impacts, str(images_dir / "04_subs.png"), dpi=dpi,
        )

    result["xg_hist"] = plot_xg_histogram(
        computed.ldi_result,
        raw.score.home, raw.score.away,
        raw.home_team.name, raw.away_team.name,
        str(images_dir / "06_xg_hist.png"),
        dpi=dpi,
    )

    # Add lineup visualization
    try:
        from src.visualizer.lineup import plot_lineup
        lineup_path = plot_lineup(raw, str(images_dir / "00_lineup.png"), dpi=dpi)
        if lineup_path:
            result["lineup"] = lineup_path
    except Exception as e:
        logger.warning(f"Lineup viz skipped: {e}")

    rel = {}
    for k, v in result.items():
        if v:
            rel[k] = str(Path(v).relative_to(output_dir)).replace("\\", "/")
        else:
            rel[k] = ""
    return rel


def _save_tactical_excel(tactical_data: dict, tactical_narrative: str,
                         home_name: str, away_name: str, output_path: str):
    """保存战术分析 Excel 文件，含所有数据及 LLM 叙事。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.warning("openpyxl not installed, skip Excel export")
        return

    wb = openpyxl.Workbook()

    # Style definitions
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    green_fill = PatternFill(start_color="2ecc71", end_color="2ecc71", fill_type="solid")
    blue_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    data_font = Font(size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header(ws, row, cols, fill):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # ── Sheet 1: 战术维度 ──
    ws1 = wb.active
    ws1.title = "战术维度"
    ws1.append([f"{home_name} 战术维度", "值", f"{away_name} 战术维度", "值", "Gap"])
    style_header(ws1, 1, 5, header_fill)

    home_raw = tactical_data["home"]["tactical_raw"]
    away_raw = tactical_data["away"]["tactical_raw"]
    home_rel = tactical_data["home"]["match_relative"]
    labels = [
        ("长传占比", "long_ball_ratio"), ("传中占比", "cross_ratio"),
        ("三区传球占比", "final_third_pass_ratio"), ("向前传球比", "forward_ratio"),
        ("每射传球", "passes_per_shot"), ("PPDA", "ppda"),
        ("高位抢断比", "high_press_ratio"), ("解围倾向", "clearance_ratio"),
    ]
    for cn, key in labels:
        gap_key = key.replace("_ratio", "_gap").replace("_per_shot", "_gap")
        ws1.append([cn, home_raw.get(key, 0), cn, away_raw.get(key, 0), home_rel.get(gap_key, "-")])
    auto_width(ws1)

    # ── Sheet 2: 执行效果 ──
    ws2 = wb.create_sheet("执行效果")
    ws2.append(["球队", "进攻/防守", "维度", "投入Gap", "ROI值", "得分", "判定"])
    style_header(ws2, 1, 7, header_fill)
    v_map = {"effective": "奏效", "fail": "失败", "undetermined": "数据不足"}
    for team_key, tname in [("home", home_name), ("away", away_name)]:
        exec_data = tactical_data[team_key]["execution"]
        for phase in ["attack", "defense"]:
            pd_data = exec_data[phase]
            for d in pd_data.get("dimensions", []):
                ws2.append([tname, phase, d["dim"], d.get("input_gap", "-"),
                            d.get("roi_value", "-"), d["score"],
                            v_map.get(pd_data["verdict"], pd_data["verdict"])])
            ws2.append([tname, phase, f"总分", "", "", pd_data["total_score"], ""])
    auto_width(ws2)

    # ── Sheet 3: 比赛走势 ──
    ws3 = wb.create_sheet("比赛走势")
    mf = tactical_data["match_flow"]

    ws3.append(["节奏主导权"])
    ws3.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws3.append(["切换次数", mf["rhythm"]["swings"]])
    ws3.append(["判定", mf["rhythm"]["verdict"]])

    ws3.append([])
    ws3.append(["PPDA 压迫强度 (全场)"])
    ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True, size=12)
    ws3.append(["球队", "全场PPDA"])
    for team_key in ["home", "away"]:
        tname = home_name if team_key == "home" else away_name
        ppda_val = mf["ppda"][team_key]["full_match"]
        ws3.append([tname, ppda_val])

    ws3.append([])
    ws3.append(["关键事件冲击"])
    ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True, size=12)
    ws3.append(["分钟", "事件类型", "球队", "触发条件", "上下文"])
    for ev in mf.get("key_event_impacts", []):
        team = home_name if ev.get("team") == "home" else away_name
        ws3.append([ev["minute"], ev["event_type"], team, ev.get("trigger", ""), ev.get("context", "")])
    auto_width(ws3)

    # ── Sheet 4: 教练博弈 ──
    ws4 = wb.create_sheet("教练博弈")
    coaching = tactical_data["coaching"]
    ws4.append(["风格碰撞", coaching["style_clash"]])
    ws4.append(["克制分", home_name, away_name])
    ws4.append(["", coaching["tactical_mismatch"]["home"], coaching["tactical_mismatch"]["away"]])
    ws4.append([])
    ws4.append(["克制对"])
    ws4.cell(row=ws4.max_row, column=1).font = Font(bold=True, size=12)
    ws4.append(["进攻方", "进攻维度", "防守维度", "结果"])
    for p in coaching.get("mismatch_pairs", []):
        ot = home_name if p.get("off_team") == "home" else away_name
        ws4.append([ot, p["off_dim"], p["def_dim"], p["result"]])
    auto_width(ws4)

    # ── Sheet 5: LLM 叙事 ──
    ws5 = wb.create_sheet("LLM叙事")
    ws5.column_dimensions["A"].width = 100
    narrative_font = Font(size=11)
    # Split by paragraphs
    paragraphs = tactical_narrative.strip().split("\n")
    for i, para in enumerate(paragraphs, 1):
        cell = ws5.cell(row=i, column=1, value=para)
        cell.font = narrative_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if para.startswith("【"):
            cell.font = Font(bold=True, size=12, color="2ecc71")

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="AI 足球比赛分析报告生成")
    parser.add_argument("--match", type=int, help="比赛 ID (fixture id)")
    parser.add_argument("--league", type=int, default=1, help="联赛 ID (默认: 1=世界杯)")
    parser.add_argument("--date", type=str, help="日期 YYYY-MM-DD")
    parser.add_argument("--dry-run", action="store_true", help="仅采集数据并计算指标")
    parser.add_argument("--no-images", action="store_true", help="跳过图表生成")
    parser.add_argument("--use-cache", action="store_true", help="使用缓存的原始数据（不调用SportMonks API）")
    parser.add_argument("--config", type=str, default="config.yaml", help="配置文件路径")
    args = parser.parse_args()

    config = load_config(args.config)

    if args.league and args.date:
        client = SportMonksClient(config["sportmonks"])
        fixtures = client.get_fixtures_by_date(
            args.league,
            config.get("competition", {}).get("season", 2026),
            args.date,
        )
        match_ids = [f["fixture"]["id"] for f in fixtures if f["fixture"]["id"]]
        logger.info(f"找到 {len(match_ids)} 场比赛: {args.date}")
    elif args.match:
        match_ids = [args.match]
    else:
        parser.error("请指定 --match 或 (--league + --date)")

    for match_id in match_ids:
        logger.info(f"{'='*60}")
        logger.info(f"处理比赛 #{match_id}...")

        try:
            if args.use_cache:
                from src.collector.api_client import load_cached_raw
                raw = load_cached_raw(match_id)
                logger.info(f"使用缓存数据: {raw.home_team.name} vs {raw.away_team.name}")
            else:
                raw = fetch_all(match_id, config["sportmonks"])
        except Exception as e:
            logger.error(f"数据采集失败 #{match_id}: {e}")
            continue

        computed = compute_all(raw)

        logger.info("Analyzing trends...")
        from src.engine.trends import analyze_trends
        trend_analysis = analyze_trends(raw)

        logger.info("Detecting signals...")
        from src.engine.signals import detect_all, get_top_signals
        all_signals = detect_all(raw, None, trend_analysis)
        top_signals = get_top_signals(all_signals, top_n=6)

        signal_names = [s.name for s in top_signals]
        logger.info(f"Top signals: {signal_names}")
        logger.info(f"  ({len(all_signals)} total signals detected, top 6 selected for LLM)")

        # --- v3: layer 1 hard facts & sub impact analysis ---
        from src.engine.cross_insights import compute_cross_insights
        from src.engine.sub_impact import analyze_sub_impacts

        logger.info("Computing hard facts (Layer 1)...")
        sub_impacts = analyze_sub_impacts(raw)
        hard_facts = compute_cross_insights(raw, sub_impacts)
        logger.info(f"Sub impacts: {len(sub_impacts)} substitutions analyzed")
        logger.info(f"Hard facts: possession_efficiency={hard_facts.possession_xg_ratio_home:.3f}/{hard_facts.possession_xg_ratio_away:.3f}")

        computed_path = Path("data/computed") / f"{match_id}.json"
        computed_path.parent.mkdir(parents=True, exist_ok=True)
        import dataclasses

        computed_dict = dataclasses.asdict(computed)
        # Save ALL detected signals (not just top 6)
        computed_dict["signals"] = [
            {"name": s.name, "category": s.category, "strength": s.strength,
             "narrative_hint": s.narrative_hint, "evidence": s.evidence}
            for s in all_signals
        ]
        computed_dict["top_signals"] = signal_names
        with open(computed_path, "w", encoding="utf-8") as f:
            json.dump(computed_dict, f, ensure_ascii=False, indent=2, default=str)

        logger.info(
            f"指标: CI({computed.home_ci}/{computed.away_ci}) "
            f"TCR({computed.home_tcr}/{computed.away_tcr}) "
            f"PE({computed.home_pe}/{computed.away_pe}) "
            f"标签: {computed.tags}"
        )

        if args.dry_run:
            logger.info(f"Dry-run 完成，数据已保存至 data/raw/{match_id}/ 和 data/computed/")
            continue

        safe_home = "".join(c if c.isalnum() or c in "-_" else "_" for c in raw.home_team.name)
        safe_away = "".join(c if c.isalnum() or c in "-_" else "_" for c in raw.away_team.name)
        output_dir = Path("output") / f"{match_id}_{safe_home}_vs_{safe_away}"

        # --- v3: six-section narrative ---
        from src.composer.data_builder import build_narrative_v3
        from src.composer.prompt_loader import PromptLoader
        from src.generator.llm_client import LLMClient

        llm = LLMClient(config["llm"])
        pl = PromptLoader("prompts")
        sys_p, user_p = build_narrative_v3(raw, computed, top_signals, hard_facts, sub_impacts, trend_analysis, pl)
        logger.info("Calling LLM for v3 narrative...")
        narrative_text = llm.generate(sys_p, user_p)
        logger.info("Narrative (v3) generated.")

        # --- tactical analysis v2 ---
        logger.info("Computing tactical analysis (四层因果模型)...")
        from src.engine.tactical_insights import compute_tactical_analysis
        tactical_data = compute_tactical_analysis(raw)
        logger.info(f"Tactical: style_clash={tactical_data['coaching']['style_clash']}")

        from src.composer.tactical_prompt import build_tactical_system_and_user
        tactical_system, tactical_user_prompt = build_tactical_system_and_user(
            tactical_data,
            raw.home_team.name, raw.away_team.name,
            raw.score.home, raw.score.away,
        )
        logger.info("Calling LLM for tactical narrative...")
        tactical_narrative = llm.generate(tactical_system, tactical_user_prompt)
        logger.info("Tactical narrative generated.")

        if not args.no_images:
            image_paths = generate_all_visuals(raw, computed, config["visual"], str(output_dir), sub_impacts=sub_impacts)
            logger.info(f"图表已生成: {len(image_paths)} 张")

            # Generate tactical charts
            try:
                from src.visualizer.tactical_charts import generate_all_tactical_charts
                tactical_image_paths_raw = generate_all_tactical_charts(
                    tactical_data, raw.home_team.name, raw.away_team.name,
                    str(output_dir / "images"), dpi=config["visual"].get("dpi", 150),
                )
                tactical_image_paths = {}
                for k, v in tactical_image_paths_raw.items():
                    tactical_image_paths[k] = str(Path(v).relative_to(output_dir)).replace("\\", "/")
                logger.info(f"战术图表已生成: {len(tactical_image_paths)} 张")
            except Exception as e:
                logger.warning(f"战术图表生成失败: {e}")
                tactical_image_paths = {}
        else:
            image_paths = {}
            tactical_image_paths = {}

        # Save tactical data JSON
        tactical_json_path = output_dir / "tactical_analysis.json"
        tactical_json_path.parent.mkdir(parents=True, exist_ok=True)
        with open(tactical_json_path, "w", encoding="utf-8") as f:
            json.dump(tactical_data, f, ensure_ascii=False, indent=2, default=str)

        # Save tactical Excel
        try:
            _save_tactical_excel(tactical_data, tactical_narrative,
                                 raw.home_team.name, raw.away_team.name,
                                 str(output_dir / "tactical_analysis.xlsx"))
            logger.info("战术分析 Excel 已保存")
        except Exception as e:
            logger.warning(f"战术 Excel 保存失败: {e}")

        # ── Pressing analysis (LLM) ──
        logger.info("Computing pressing analysis...")
        pressing_narrative = ""
        try:
            from src.composer.pressing_prompt import build_pressing_prompt

            # Build goal events (format expected by pressing prompt)
            goal_events = []
            for ev in raw.events:
                if ev.event_type == "Goal" and ev.detail not in ("pen_shootout_goal", "pen_shootout_miss"):
                    team_label = "home" if ev.team_id == raw.home_team.id else "away"
                    goal_events.append({
                        "minute": ev.time_elapsed or 0,
                        "label": ev.player_name or "",
                        "team": team_label,
                    })

            # Build def_actions from trends
            windows = [(0, 15), (15, 30), (30, 45), (45, 60), (60, 75), (75, 90)]
            def _cum_at(pts, minute):
                val = 0.0
                for p in sorted(pts, key=lambda x: x.minute):
                    if p.minute <= minute:
                        val = p.value
                    else:
                        break
                return val

            def _window_actions(own_team_id: str) -> dict:
                result = {"tackles": [], "interceptions": [], "fouls": []}
                for metric, tid in [("tackles", "78"), ("interceptions", "100"), ("fouls", "56")]:
                    pts = raw.trends.get(own_team_id, {}).get(tid, [])
                    for start, end in windows:
                        delta = _cum_at(pts, end) - _cum_at(pts, start)
                        result[metric].append(round(delta, 1))
                return result

            def_actions = {
                "home": _window_actions(str(raw.home_team.id)),
                "away": _window_actions(str(raw.away_team.id)),
            }

            pressing_system, pressing_user = build_pressing_prompt(
                raw.home_team.name, raw.away_team.name,
                tactical_data["match_flow"]["ppda_trend"],
                tactical_data["match_flow"]["possession_trend"],
                tactical_data["match_flow"]["shot_segments"],
                def_actions, goal_events,
                raw.score.home, raw.score.away,
            )
            logger.info("Calling LLM for pressing narrative...")
            pressing_narrative = llm.generate(pressing_system, pressing_user)
            logger.info("Pressing narrative generated.")
        except Exception as e:
            logger.warning(f"Pressing analysis failed: {e}")

        # ── 环节一：战术文章取名 ──
        logger.info("Generating article titles...")
        article_titles = []
        try:
            from src.composer.article_naming_prompt import build_article_naming_prompt, parse_article_titles

            # Build match context
            stage = raw.stage_info or {}
            league_name = stage.get("name", "")
            stage_text = f"赛事：{league_name}" if league_name else ""
            has_pen = raw.score.penalty_home is not None and raw.score.penalty_home > 0
            ctx_lines = [stage_text] if stage_text else []
            ctx_lines.append(f"对阵：{raw.home_team.name} vs {raw.away_team.name}")
            if has_pen:
                reg_home = raw.score.home - (raw.score.penalty_home or 0)
                reg_away = raw.score.away - (raw.score.penalty_away or 0)
                ctx_lines.append(f"最终比分：{raw.score.home}-{raw.score.away}（常规时间 {reg_home}-{reg_away}，点球 {raw.score.penalty_home}-{raw.score.penalty_away}）")
            else:
                ctx_lines.append(f"比分：{raw.home_team.name} {raw.score.home}-{raw.score.away} {raw.away_team.name}")

            # Key events text
            key_ev_lines = []
            for ev in raw.events:
                if ev.event_type in ("Goal", "Card") and ev.detail not in ("pen_shootout_goal", "pen_shootout_miss"):
                    icon = "⚽" if ev.event_type == "Goal" else "🟨"
                    desc = f"{icon} {ev.time_elapsed}' {ev.player_name or ''}"
                    if ev.event_type == "Goal" and ev.assist_name:
                        desc += f"（助攻：{ev.assist_name}）"
                    key_ev_lines.append(desc)
            key_events_text = "\n".join(key_ev_lines[:30])

            naming_sys, naming_user = build_article_naming_prompt(
                match_context="\n".join(ctx_lines),
                tactical_summary=tactical_narrative[:2000],
                pressing_summary=pressing_narrative[:1500] if pressing_narrative else tactical_narrative[:1500],
                key_events=key_events_text,
            )
            logger.info("Calling LLM for article titles...")
            naming_text = llm.generate(naming_sys, naming_user, max_tokens=2048)
            article_titles = parse_article_titles(naming_text)
            logger.info(f"Article titles generated: {len(article_titles)} titles")

            # Save article titles JSON
            titles_data = {
                "match_id": match_id,
                "match_context": "\n".join(ctx_lines),
                "titles": article_titles,
            }
            titles_path = output_dir / "article_titles.json"
            with open(titles_path, "w", encoding="utf-8") as f:
                json.dump(titles_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.warning(f"Article naming failed: {e}")

        # ── 环节二：球员人物特稿推荐 ──
        logger.info("Generating player feature recommendations...")
        player_features_md = ""
        try:
            from src.composer.player_feature_prompt import build_player_feature_prompt
            from src.engine.player_feature_selector import select_candidates

            # Build player summary (reuse existing player_analysis prompt data if available)
            # For now, build basic player data from raw
            hname = raw.home_team.name
            aname = raw.away_team.name
            home_logo = raw.home_team.logo_url
            away_logo = raw.away_team.logo_url

            # Build match context
            stage = raw.stage_info or {}
            league_name = stage.get("name", "")
            match_ctx = "\n".join(ctx_lines) if ctx_lines else f"对阵：{raw.home_team.name} vs {raw.away_team.name}，比分 {raw.score.home}-{raw.score.away}"

            # Build player data list for candidate selection
            players_data = []
            for p in raw.home_players:
                if p.minutes_played <= 0:
                    continue
                players_data.append({
                    "name": p.name, "team": hname,
                    "position": p.position, "minutes": p.minutes_played,
                    "rating": p.rating or 0, "goals": p.goals, "assists": p.assists,
                    "shots_total": p.shots_total, "shots_on": p.shots_on,
                    "xg": p.xg, "xgot": p.xgot,
                    "passes_total": p.passes_total, "passes_accuracy": p.passes_accuracy,
                    "passes_key": p.passes_key, "passes_final_third": p.passes_final_third,
                    "tackles_total": p.tackles_total, "tackles_interceptions": p.tackles_interceptions,
                    "duels_won": p.duels_won, "duels_total": p.duels_total,
                    "dribbles_success": p.dribbles_success, "dribbles_attempts": p.dribbles_attempts,
                    "saves": p.saves, "saves_inside_box": p.saves_inside_box,
                    "error_lead_to_goal": p.error_lead_to_goal,
                    "photo_url": p.photo_url, "man_of_match": p.man_of_match,
                    "fouls_drawn": p.fouls_drawn, "fouls_committed": p.fouls_committed,
                    "ball_recoveries": p.ball_recoveries,
                    "yellowcards": p.yellowcards, "redcards": p.redcards,
                })
            for p in raw.away_players:
                if p.minutes_played <= 0:
                    continue
                players_data.append({
                    "name": p.name, "team": aname,
                    "position": p.position, "minutes": p.minutes_played,
                    "rating": p.rating or 0, "goals": p.goals, "assists": p.assists,
                    "shots_total": p.shots_total, "shots_on": p.shots_on,
                    "xg": p.xg, "xgot": p.xgot,
                    "passes_total": p.passes_total, "passes_accuracy": p.passes_accuracy,
                    "passes_key": p.passes_key, "passes_final_third": p.passes_final_third,
                    "tackles_total": p.tackles_total, "tackles_interceptions": p.tackles_interceptions,
                    "duels_won": p.duels_won, "duels_total": p.duels_total,
                    "dribbles_success": p.dribbles_success, "dribbles_attempts": p.dribbles_attempts,
                    "saves": p.saves, "saves_inside_box": p.saves_inside_box,
                    "error_lead_to_goal": p.error_lead_to_goal,
                    "photo_url": p.photo_url, "man_of_match": p.man_of_match,
                    "fouls_drawn": p.fouls_drawn, "fouls_committed": p.fouls_committed,
                    "ball_recoveries": p.ball_recoveries,
                    "yellowcards": p.yellowcards, "redcards": p.redcards,
                })

            # Build raw events list for key events detection
            raw_events_list = []
            for ev in raw.events:
                raw_events_list.append({
                    "player_name": ev.player_name,
                    "event_type": ev.event_type,
                    "detail": ev.detail,
                    "minute": ev.time_elapsed,
                    "team_id": ev.team_id,
                    "assist_name": ev.assist_name,
                })

            # Detect key events
            from src.engine.key_events import detect_key_events
            goal_events_raw = [e for e in raw_events_list if e["event_type"] in ("Goal", "goal")]
            sub_events_raw = [e for e in raw_events_list if e["event_type"] in ("subst", "substitution")]
            key_events_result = detect_key_events(
                goal_events_raw, sub_events_raw,
                raw.home_team.id, raw.away_team.id,
                raw.score.home, raw.score.away,
            )

            # Run player insights for tags
            # Build lineups for detector
            from build_player_analysis import player_to_lineup, parse_events as pipeline_parse_events, load_or_fetch
            RAW_dict = load_or_fetch(match_id)
            lineups_list = [player_to_lineup(p, raw.home_team.id) for p in RAW_dict.get('home_players', [])]
            lineups_list += [player_to_lineup(p, raw.away_team.id) for p in RAW_dict.get('away_players', [])]
            events_parsed = pipeline_parse_events(RAW_dict)
            max_min = max(p.get('minutes_played', 0) or 0
                          for p in RAW_dict.get('home_players', []) + RAW_dict.get('away_players', []))
            end_min = max(90, max_min) if max_min > 0 else 90

            from src.engine.player_insights import run_all_detectors
            detector_results = run_all_detectors(
                lineups_list, raw.home_team.id, raw.away_team.id,
                raw.score.home, raw.score.away,
                events_parsed, end_min,
                home_name=raw.home_team.name, away_name=raw.away_team.name,
            )

            # Build summaries dict from player data (use raw ratings as fallback)
            player_summaries = {}
            for p in raw.home_players + raw.away_players:
                if p.minutes_played > 0:
                    summary_parts = []
                    if p.goals:
                        summary_parts.append(f"打入{p.goals}球")
                    if p.assists:
                        summary_parts.append(f"助攻{p.assists}次")
                    if p.shots_on:
                        summary_parts.append(f"射正{p.shots_on}次")
                    if p.passes_key:
                        summary_parts.append(f"关键传球{p.passes_key}次")
                    if p.saves and p.saves >= 3:
                        summary_parts.append(f"完成{p.saves}次扑救")
                    player_summaries[p.name] = "，".join(summary_parts) if summary_parts else f"出场{p.minutes_played}分钟，评分{p.rating or '-'}"

            # Select candidates
            candidates = select_candidates(
                players_data, detector_results, key_events_result,
                raw_events_list, player_summaries, max_candidates=8,
            )
            logger.info(f"Player feature candidates: {len(candidates)} selected")

            # Fetch news for candidate players' teams
            try:
                if config.get("sportmonks", {}).get("api_token"):
                    from src.collector.api_client import SportMonksClient
                    sm_client = SportMonksClient(config["sportmonks"])
                    team_ids = list(set(
                        raw.home_team.id for c in candidates
                    ) | set(raw.away_team.id for c in candidates))
                    all_news = sm_client.get_newsfeeds(team_ids, limit=20)
                    # Match news to candidates by name keyword
                    for c in candidates:
                        c_news = []
                        for n in all_news[:30]:
                            content = (n.get("title", "") + " " + n.get("content", ""))
                            if c.name.lower() in content.lower() and len(c_news) < 3:
                                c_news.append(f"({n.get('created_at', '?'):10s}) {n['title'][:80]}")
                        c.news = c_news
                    logger.info(f"News fetched: {len(all_news)} articles")
            except Exception as e:
                logger.warning(f"News fetch failed: {e}")

            # Build LLM prompt and generate
            stage_text = f"赛事：{league_name}" if league_name else "赛事信息暂缺"
            feature_sys, feature_user = build_player_feature_prompt(
                match_context=match_ctx,
                stage_info=stage_text,
                tactical_summary=tactical_narrative[:1500],
                pressing_summary=pressing_narrative[:1000] if pressing_narrative else tactical_narrative[:1000],
                candidates=candidates,
            )
            logger.info("Calling LLM for player feature recommendations...")
            player_features_md = llm.generate(feature_sys, feature_user, max_tokens=4800)
            logger.info(f"Player features generated: {len(player_features_md)} chars")

            # Save player features Markdown
            features_path = output_dir / "player_features.md"
            with open(features_path, "w", encoding="utf-8") as f:
                f.write(player_features_md)
            logger.info(f"Player features saved to {features_path}")
        except Exception as e:
            logger.warning(f"Player feature recommendation failed: {e}")
            import traceback
            traceback.print_exc()

        from src.reporter.build_report import build_report_v3_html
        report_path = build_report_v3_html(
            raw, narrative_text, image_paths, str(output_dir),
            hard_facts=hard_facts, sub_impacts=sub_impacts,
            signals=all_signals, computed=computed,
            tactical_narrative=tactical_narrative,
            tactical_data=tactical_data,
            tactical_image_paths=tactical_image_paths,
            article_titles=article_titles,
            player_features_md=player_features_md,
            pressing_narrative=pressing_narrative,
        )
        logger.info(f"报告 (v3 HTML) 已生成: {report_path}")

    logger.info(f"\n完成！共处理 {len(match_ids)} 场比赛。")


if __name__ == "__main__":
    main()
