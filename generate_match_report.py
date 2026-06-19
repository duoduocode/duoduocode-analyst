"""
统一比赛报告生成器 — 战术报告 V2 + 球员贡献检测器 V6

一次性生成:
  1. 战术分析报告 V2 (HTML + JSON + Excel + 图表)
  2. 球员贡献检测器 V6 (JSON + Excel + 球员卡片)

用法:
  python generate_match_report.py 19683241
  python generate_match_report.py 19683241 --no-llm          # 跳过 LLM 叙事
  python generate_match_report.py 19683241 --cards-only       # 仅生成球员卡片
  python generate_match_report.py 19683241 --tactical-only    # 仅生成战术报告
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import time
from pathlib import Path

import requests
try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

THIS_DIR = os.path.dirname(os.path.abspath(__file__))


# ═══════════════════════════════════════════════════════════════
# 配置加载
# ═══════════════════════════════════════════════════════════════

def load_config():
    import yaml
    with open("config.yaml", "r", encoding="utf-8") as f:
        raw = f.read()
    for key, value in os.environ.items():
        raw = raw.replace("${" + key + "}", value)
    return yaml.safe_load(raw)


# ═══════════════════════════════════════════════════════════════
# 数据加载
# ═══════════════════════════════════════════════════════════════

def load_match_data(match_id: int, config: dict):
    """加载比赛数据（优先缓存）。"""
    from src.collector.api_client import load_cached_raw, fetch_all

    os.makedirs("data/raw", exist_ok=True)
    try:
        raw = load_cached_raw(match_id)
        logger.info(f"使用缓存: data/raw/{match_id}/raw_data.json")
        return raw
    except Exception:
        logger.info(f"缓存不可用，通过 API 拉取...")
        return fetch_all(match_id, config["sportmonks"])


# ═══════════════════════════════════════════════════════════════
# 最终比分计算（常规 + 加时 + 点球）
# ═══════════════════════════════════════════════════════════════

def compute_full_score(raw) -> tuple:
    """返回 (total_home, total_away, pen_home, pen_away, has_penalties)."""
    score = raw.score
    home_id = raw.home_team.id
    pen_home = pen_away = 0
    for e in raw.events:
        if e.event_type == "Goal" and e.detail == "pen_shootout_goal":
            if e.team_id == home_id:
                pen_home += 1
            else:
                pen_away += 1
    total_home = score.home + (score.extratime_home or 0) + pen_home
    total_away = score.away + (score.extratime_away or 0) + pen_away
    return total_home, total_away, pen_home, pen_away, (pen_home > 0 or pen_away > 0)


# ═══════════════════════════════════════════════════════════════
# 管道 A: 战术报告 V2
# ═══════════════════════════════════════════════════════════════

def generate_tactical_report_v2(raw, config: dict, output_dir: Path):
    """生成战术分析报告 V2 (HTML + JSON + Excel + 图表)。"""
    logger.info("=" * 40)
    logger.info("管道 A: 战术报告 V2")

    home_name = raw.home_team.name
    away_name = raw.away_team.name
    score = raw.score
    total_home, total_away, pen_home, pen_away, has_penalties = compute_full_score(raw)

    images_dir = output_dir / "images"
    images_dir.mkdir(parents=True, exist_ok=True)

    # ── 1. 战术分析计算 ──
    logger.info("计算战术分析 (四层因果模型)...")
    from src.engine.tactical_insights import compute_tactical_analysis
    tactical_data = compute_tactical_analysis(raw)
    logger.info(f"  风格碰撞: {tactical_data['coaching']['style_clash']}")

    # ── 1.5 比赛过程概述 + 新闻摘要 + 图片 (豆包联网搜索 v4) ──
    match_overview = ""
    match_overview_sys = ""
    match_overview_user = ""
    stage_name = (raw.stage_info or {}).get("name", "")
    try:
        from src.generator.llm_client import DoubaoClient, LLMClient

        llm = LLMClient(config["llm"])
        doubao = DoubaoClient(config.get("doubao", {}))

        stage_year = ""
        start_date = (raw.stage_info or {}).get("starting_at", "")
        if start_date:
            stage_year = start_date[:4]

        # 去中文名
        _NAME_MAP_DB = {
            "netherlands": "荷兰", "japan": "日本", "england": "英格兰",
            "france": "法国", "germany": "德国", "spain": "西班牙",
            "italy": "意大利", "portugal": "葡萄牙", "argentina": "阿根廷",
            "brazil": "巴西", "sweden": "瑞典", "tunisia": "突尼斯",
            "korea": "韩国", "south korea": "韩国",
        }
        home_cn = _NAME_MAP_DB.get(home_name.lower().strip(), home_name)
        away_cn = _NAME_MAP_DB.get(away_name.lower().strip(), away_name)

        # ─── 三轮豆包联网搜索 ───
        prompts = {
            "pre": f"请联网搜索{stage_year}年美加墨世界杯{home_cn}对{away_cn}的赛前新闻（赛前阵容、伤病、前瞻、历史交锋等）。\n按格式输出：## 赛前新闻列表（含来源）\n## 赛前新闻摘要(160-300字)",
            "match": f"请联网搜索{stage_year}年美加墨世界杯{home_cn}对{away_cn}的比赛战报（首发、进球事件、关键数据、赛后评价）。\n按格式输出：## 比赛基本信息+关键事件时间线+双方首发阵容\n## 比赛赛况摘要(380-500字)",
            "post": f"请联网搜索{stage_year}年美加墨世界杯{home_cn}对{away_cn}的赛后新闻（球员评价、纪录、出线形势等）。\n按格式输出：## 赛后新闻列表（含来源）\n## 赛后新闻摘要(160-300字)",
        }

        all_article_urls = []
        all_summaries = {}
        web_dir = output_dir / "web_context"
        web_dir.mkdir(parents=True, exist_ok=True)

        for mode, label in [("pre", "赛前"), ("match", "赛况"), ("post", "赛后")]:
            logger.info(f"  豆包联网搜索 [{label}]...")
            try:
                t0 = time.time()
                result = doubao.search(prompts[mode])
                elapsed = time.time() - t0
                all_summaries[mode] = result["content"]
                all_article_urls.extend(result["article_urls"])
                # 保存文本
                txt_path = web_dir / f"{mode}.txt"
                with open(txt_path, "w", encoding="utf-8") as f:
                    f.write(f"# {label}新闻摘要\n> 豆包联网搜索 | {elapsed:.1f}s\n> tokens: in={result['input_tokens']} out={result['output_tokens']}\n\n")
                    f.write(result["content"])
                logger.info(f"    {label}.txt: {len(result['content'])} 字符 | {len(result['article_urls'])} URLs")
            except Exception as e:
                logger.warning(f"    豆包 [{label}] 搜索失败: {e}")
                all_summaries[mode] = f"（搜索失败: {e}）"
            time.sleep(2)

        # ─── 去重文章URL ───
        unique_urls = list(dict.fromkeys(all_article_urls))
        logger.info(f"  去重后文章URL: {len(unique_urls)} 条")

        # ─── 从文章页提取图片 ───
        img_dir = web_dir / "images"
        img_dir.mkdir(parents=True, exist_ok=True)

        def _extract_images_from_page(url: str) -> list:
            """从文章页提取比赛相关图片。
            优先队名/关键词匹配，无匹配时宽模式兜底（文章来自豆包搜索，页面本身相关）。
            """
            if BeautifulSoup is None:
                return []
            try:
                r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"},
                                timeout=15, allow_redirects=True, verify=False)
                r.raise_for_status()
            except Exception:
                return []
            soup = BeautifulSoup(r.text, "html.parser")
            container = soup.find("article") or soup.find("main") or soup.find("body")
            if not container:
                return []

            skip_d = ("beacon", "tracking", "pixel", "doubleclick", "analytics")
            skip_p = ("logo", "icon", "avatar", "qr", "ewm", "share", "arrow", "btn",
                      "close", "weixin", "wechat", "code", "homepage", "default",
                      "fileftp", "login", "placeholder", "blank",
                      "/user/", "discusser", "bg@", "top-video", "inside-top",
                      "style/", "static.", "favicon", "emoticon")
            # 队名关键词
            team_kw = set()
            for name in (home_name, away_name, home_cn, away_cn):
                if name and len(name) > 2:
                    team_kw.add(name.lower().strip())
            # 足球通用关键词
            match_kw = {"世界杯", "world cup", "进球", "goal", "庆祝", "celebrat",
                        "首发", "lineup", "2026", "fifa"}
            # 加入Top 6球员名
            for p in sorted(raw.home_players + raw.away_players,
                           key=lambda x: x.rating or 0, reverse=True)[:6]:
                name = (p.name or "").strip()
                if name:
                    match_kw.add(name.lower())
                    parts = name.split()
                    if len(parts) > 1:
                        match_kw.add(parts[-1].lower())

            # ─── 全量扫描，记录"候选图片" ───
            all_candidates = []  # 所有通过尺寸/后缀过滤的图片
            seen = set()

            for img in container.find_all("img"):
                src = img.get("src") or img.get("data-src") or img.get("data-original") or img.get("data-lazy-src") or ""
                if not src:
                    continue
                if src.startswith("//"):
                    src = "https:" + src
                if not src.startswith("http") or src in seen:
                    continue
                seen.add(src)
                sl = src.lower()
                if any(k in sl for k in skip_d):
                    continue
                if any(k in sl for k in skip_p):
                    continue
                if not any(sl.endswith(e) for e in (".jpg", ".jpeg", ".png", ".webp")):
                    continue
                sz = re.search(r'[_-](\d+)x(\d+)', src)
                if sz and (int(sz.group(1)) < 100 or int(sz.group(2)) < 100):
                    continue
                alt = (img.get("alt") or "").strip()
                alt_l = alt.lower()
                info = {"url": src, "alt": alt, "source_url": url}

                # ─── 黑名单: alt/URL 明确不是比赛图片的 ───
                alt_blacklist = ("avatar", "header", "discusser", "bg@", "bg-",
                                 "favicon", "emoticon", "logo", "icon-",
                                 "top-video", "inside-top", "background")
                if any(kw in alt_l for kw in alt_blacklist):
                    continue  # 直接丢弃

                # 标记匹配级别
                matched = False
                # Level 1: alt/URL 含队名
                if any(kw in alt_l for kw in team_kw) or any(kw in sl for kw in team_kw):
                    info["level"] = 1
                    matched = True
                # Level 2: alt/URL 含通用关键词
                elif any(kw in alt_l for kw in match_kw) or any(kw in sl for kw in match_kw):
                    info["level"] = 2
                    matched = True
                # Level 3: 父元素文本含队名
                else:
                    parent_text = ""
                    for p in img.parents:
                        txt = p.get_text(strip=True)
                        if len(txt) > 10:
                            parent_text = txt; break
                    if parent_text and team_kw:
                        pl = parent_text.lower()
                        if any(kw in pl for kw in team_kw):
                            info["level"] = 3
                            matched = True

                if matched:
                    all_candidates.append(info)
                else:
                    info["level"] = 4
                    all_candidates.append(info)  # 无匹配，但先保留

            # ─── 排序: level小的(匹配度高)优先；level=4的放到最后 ───
            all_candidates.sort(key=lambda x: x.get("level", 99))
            # 去重
            seen2 = set()
            final = []
            for c in all_candidates:
                if c["url"] not in seen2:
                    seen2.add(c["url"])
                    final.append(c)
            # 最多60张候选，后续下载按字节大小再过滤
            return final[:60]

        all_images = []
        for i, url in enumerate(unique_urls):
            imgs = _extract_images_from_page(url)
            all_images.extend(imgs)
            time.sleep(0.3)

        # 图片去重
        seen_urls = set()
        unique_imgs = []
        for img in all_images:
            if img["url"] not in seen_urls:
                seen_urls.add(img["url"])
                unique_imgs.append(img)
        logger.info(f"  图片: 原始{len(all_images)}张 → 去重{len(unique_imgs)}张")

        # ─── 下载图片 ───
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        downloaded = []
        for i, img in enumerate(unique_imgs):
            url = img["url"]
            try:
                r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"},
                                timeout=15, verify=False)
                r.raise_for_status()
                if len(r.content) < 5 * 1024:
                    continue
                ext = ".jpg"
                for e in (".jpg", ".jpeg", ".png", ".webp"):
                    if e in url.lower():
                        ext = e; break
                fname = f"img_{len(downloaded)+1:03d}{ext}"
                with open(img_dir / fname, "wb") as f:
                    f.write(r.content)
                downloaded.append({"filename": fname, "url": url,
                                   "alt": img["alt"], "source_url": img["source_url"],
                                   "size_kb": len(r.content) // 1024})
            except Exception:
                pass
            time.sleep(0.15)
            if len(downloaded) >= 30:
                break
        logger.info(f"  下载成功: {len(downloaded)} 张 | 总大小: {sum(d['size_kb'] for d in downloaded)//1024} MB")

        # 保存图片索引
        with open(web_dir / "images.json", "w", encoding="utf-8") as f:
            json.dump({"match": f"{home_name} vs {away_name}", "match_id": raw.match_id,
                       "total": len(downloaded), "images": downloaded,
                       "source_articles": unique_urls,
                       "search_provider": "doubao"}, f, ensure_ascii=False, indent=2)

        # ─── 组装 match_overview ───
        # 构建关键事件时间线（SportMonks 数据）
        key_ev_lines = []
        for ev in raw.events:
            if ev.event_type in ("Goal", "Card") and ev.detail not in ("pen_shootout_goal", "pen_shootout_miss"):
                icon = "进球" if ev.event_type == "Goal" else "黄牌"
                desc = f"{icon} {ev.time_elapsed}' {ev.player_name or ''}"
                if ev.event_type == "Goal" and ev.assist_name:
                    desc += f"（助攻：{ev.assist_name}）"
                key_ev_lines.append(desc)
        key_ev_str = "\n".join(key_ev_lines[:30])

        # 构建比赛基本信息
        score_line = f"{home_name} {total_home}-{total_away} {away_name}"
        if has_penalties:
            score_line += f"（常规 {total_home-pen_home}-{total_away-pen_away}，点球 {pen_home}-{pen_away}）"
        match_info = f"赛事：{stage_year}年世界杯 {stage_name}\n对阵：{home_name} vs {away_name}\n比分：{score_line}"

        # 赛况摘要直接用作 match_overview
        match_content = all_summaries.get("match", "")
        if match_content and "（搜索失败" not in match_content:
            # 提取"赛况摘要"部分
            m = re.search(r'比赛赛况摘要.*?\n(.*)', match_content, re.DOTALL)
            if m:
                match_overview = m.group(1).strip()
            else:
                match_overview = match_content[:1500]
        if not match_overview:
            match_overview = f"{match_info}\n{key_ev_str}"

        # DeepSeek 精细润色
        try:
            from src.composer.match_overview_prompt import build_match_overview_prompt
            ov_sys, ov_user = build_match_overview_prompt(
                match_context=match_info,
                key_events_text=key_ev_str,
                news_text=match_overview[:2500],
            )
            match_overview_sys = ov_sys
            match_overview_user = ov_user
            logger.info("  DeepSeek 精细润色 match_overview...")
            for ov_attempt in range(2):
                refined = llm.generate(ov_sys, ov_user, max_tokens=600)
                if refined:
                    match_overview = refined
                    break
        except Exception as e:
            logger.warning(f"  match_overview DeepSeek 润色跳过: {e}")

        logger.info(f"  match_overview 概述长度: {len(match_overview)} 字符")

    except Exception as e:
        logger.warning(f"比赛过程概述 (豆包v4) 跳过: {e}")

    # ── 2. LLM 战术叙事 ──
    tactical_narrative = ""
    one_liner = ""
    try:
        from src.composer.tactical_prompt import build_tactical_system_and_user
        from src.generator.llm_client import LLMClient

        llm = LLMClient(config["llm"])
        tactical_system, tactical_user_prompt = build_tactical_system_and_user(
            tactical_data, home_name, away_name,
            total_home, total_away,
            pen_home=pen_home, pen_away=pen_away, stage_name=stage_name,
            match_overview=match_overview,
        )
        logger.info("调用 LLM 生成战术叙事...")
        tactical_narrative = llm.generate(tactical_system, tactical_user_prompt)
        logger.info(f"  叙事长度: {len(tactical_narrative)} 字符")

        # 提取一句话总结
        ol_match = re.search(r'【一句话总结】\s*\n(.*?)(?=\n【|\Z)', tactical_narrative, re.DOTALL)
        if ol_match:
            one_liner = ol_match.group(1).strip()
            tactical_narrative = tactical_narrative.replace(ol_match.group(0), "").strip()
    except Exception as e:
        logger.warning(f"LLM 叙事跳过: {e}")
        one_liner = f"{home_name} {total_home}-{total_away} {away_name}"

    # ── 3. 战术图表 ──
    logger.info("生成战术图表...")
    try:
        from src.visualizer.tactical_charts import generate_all_tactical_charts
        dpi = config.get("visual", {}).get("dpi", 150)
        tactical_image_paths_raw = generate_all_tactical_charts(
            tactical_data, home_name, away_name, str(images_dir), dpi=dpi,
        )
        tactical_image_paths = {}
        for k, v in tactical_image_paths_raw.items():
            tactical_image_paths[k] = str(Path(v).relative_to(output_dir)).replace("\\", "/")
        logger.info(f"  已生成: {list(tactical_image_paths.keys())}")
    except Exception as e:
        logger.warning(f"战术图表生成失败: {e}")
        tactical_image_paths = {}

    # ── 3.5 压迫分析图表 ──
    def_actions = {}
    goal_events = []
    try:
        from src.visualizer.tactical_charts import plot_pressing_effectiveness, plot_pressing_efficiency
        # Goal events
        goal_events = []
        for e in raw.events:
            if e.event_type == "Goal" and e.detail not in ("pen_shootout_goal", "pen_shootout_miss"):
                team_label = "home" if e.team_id == raw.home_team.id else "away"
                goal_events.append({
                    "minute": e.time_elapsed,
                    "label": e.player_name or "",
                    "team": team_label,
                })
        # Defensive actions per window
        windows = [(0, 15), (15, 30), (30, 45), (45, 60), (60, 75), (75, 90)]
        def _cum_at(pts, minute):
            val = 0.0
            for p in sorted(pts, key=lambda x: x.minute):
                if p.minute <= minute:
                    val = p.value
                else:
                    break
            return val

        def _window_actions(own_team_id: str) -> dict:
            result = {"tackles": [], "interceptions": [], "fouls": []}
            for metric, tid in [("tackles", "78"), ("interceptions", "100"), ("fouls", "56")]:
                pts = raw.trends.get(own_team_id, {}).get(tid, [])
                for start, end in windows:
                    delta = _cum_at(pts, end) - _cum_at(pts, start)
                    result[metric].append(round(delta, 1))
            return result

        def_actions = {
            "home": _window_actions(str(raw.home_team.id)),
            "away": _window_actions(str(raw.away_team.id)),
        }

        ppda_trend = tactical_data["match_flow"]["ppda_trend"]
        eff_path = str(images_dir / "pressing_effectiveness.png")
        plot_pressing_effectiveness(home_name, away_name,
                                    ppda_trend, tactical_data["match_flow"]["possession_trend"],
                                    tactical_data["match_flow"]["shot_segments"],
                                    goal_events, eff_path, dpi=dpi)
        tactical_image_paths["pressing_effectiveness"] = str(Path(eff_path).relative_to(output_dir)).replace("\\", "/")

        eff2_path = str(images_dir / "pressing_efficiency.png")
        plot_pressing_efficiency(home_name, away_name, ppda_trend, def_actions,
                                 eff2_path, dpi=dpi)
        tactical_image_paths["pressing_efficiency"] = str(Path(eff2_path).relative_to(output_dir)).replace("\\", "/")
        logger.info("  压迫分析图表已生成")
    except Exception as e:
        logger.warning(f"压迫分析图表跳过: {e}")

    # ── 3.6 压迫分析 LLM 叙事 ──
    pressing_narrative = ""
    try:
        if def_actions:
            from src.composer.pressing_prompt import build_pressing_prompt
            from src.generator.llm_client import LLMClient
            from src.composer.prompt_loader import PromptLoader

            llm = LLMClient(config["llm"])
            pl = PromptLoader()
            pressing_system, pressing_user = build_pressing_prompt(
                home_name, away_name,
                tactical_data["match_flow"]["ppda_trend"],
                tactical_data["match_flow"]["possession_trend"],
                tactical_data["match_flow"]["shot_segments"],
                def_actions,
                goal_events,
                total_home, total_away,
                loader=pl,
                match_overview=match_overview,
            )
            logger.info("调用 LLM 生成压迫分析叙事...")
            pressing_narrative = llm.generate(pressing_system, pressing_user)
            logger.info(f"  压迫叙事长度: {len(pressing_narrative)} 字符")
    except Exception as e:
        logger.warning(f"压迫叙事 LLM 跳过: {e}")

    # ── 4. 事件时间轴 ──
    timeline_html = ""
    timeline_png_rel = None
    try:
        from src.visualizer.tactical_charts import generate_event_timeline_html, save_timeline_png
        timeline_html = generate_event_timeline_html(raw, home_name, away_name)
        timeline_png_path = str(images_dir / "timeline.png")
        save_timeline_png(raw, home_name, away_name, timeline_png_path)
        timeline_png_rel = str(Path(timeline_png_path).relative_to(output_dir)).replace("\\", "/")
        logger.info("  事件时间轴已生成")
    except Exception as e:
        logger.warning(f"事件时间轴跳过: {e}")

    # ── 5. 阵容图 ──
    lineup_html = ""
    try:
        from src.visualizer.lineup import generate_lineup_html, save_lineup_png
        lineup_html = generate_lineup_html(raw)
        lineup_png_path = str(images_dir / "lineup.png")
        save_lineup_png(raw, lineup_png_path)
        logger.info("  阵容图已生成")
    except Exception as e:
        logger.warning(f"阵容图跳过: {e}")

    # ── 环节一：文章取名 ──
    article_titles = []
    naming_text = ""
    try:
        from src.composer.article_naming_prompt import build_article_naming_prompt, parse_article_titles
        from src.generator.llm_client import LLMClient

        llm = LLMClient(config["llm"])
        # Build match context
        ctx_lines = []
        if stage_name:
            ctx_lines.append(f"赛事：{stage_name}")
        ctx_lines.append(f"对阵：{home_name} vs {away_name}")
        if has_penalties:
            reg_home = total_home - pen_home
            reg_away = total_away - pen_away
            ctx_lines.append(f"最终比分：{total_home}-{total_away}（常规+加时 {reg_home}-{reg_away}，点球 {pen_home}-{pen_away}）")
        else:
            ctx_lines.append(f"比分：{home_name} {total_home}-{total_away} {away_name}")
        match_ctx = "\n".join(ctx_lines)

        # Key events text
        key_ev_lines = []
        for ev in raw.events:
            if ev.event_type in ("Goal", "Card") and ev.detail not in ("pen_shootout_goal", "pen_shootout_miss"):
                icon = "⚽" if ev.event_type == "Goal" else "🟨"
                desc = f"{icon} {ev.time_elapsed}' {ev.player_name or ''}"
                if ev.event_type == "Goal" and ev.assist_name:
                    desc += f"（助攻：{ev.assist_name}）"
                key_ev_lines.append(desc)
        key_events_text = "\n".join(key_ev_lines[:30])

        naming_sys, naming_user = build_article_naming_prompt(
            match_context=match_ctx,
            tactical_summary=tactical_narrative[:2000],
            pressing_summary=pressing_narrative[:1500] if pressing_narrative else tactical_narrative[:1500],
            key_events=key_events_text,
            match_overview=match_overview,
        )
        logger.info("调用 LLM 生成文章标题...")
        naming_text = llm.generate(naming_sys, naming_user, max_tokens=2048)
        article_titles = parse_article_titles(naming_text)
        # Fallback: if structured parser returns 0, extract any 《title》 patterns
        if not article_titles and naming_text:
            fallback = _fallback_parse_titles(naming_text)
            if fallback:
                article_titles = fallback
                logger.info(f"  使用 fallback 解析器恢复 {len(article_titles)} 个标题")
        logger.info(f"  生成 {len(article_titles)} 个标题")

        # Save article titles JSON
        titles_data = {
            "match_id": raw.match_id,
            "match_context": match_ctx,
            "titles": article_titles,
        }
        titles_path = output_dir / "article_titles.json"
        with open(titles_path, "w", encoding="utf-8") as f:
            json.dump(titles_data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"文章取名跳过: {e}")

    # ── 环节二：球员人物特稿推荐 ──
    player_features_md = ""
    try:
        from src.composer.player_feature_prompt import build_player_feature_prompt
        from src.engine.player_feature_selector import select_candidates
        from src.generator.llm_client import LLMClient

        llm = LLMClient(config["llm"])

        # Build players_data
        players_data = []
        for p in raw.home_players:
            if p.minutes_played <= 0:
                continue
            players_data.append({
                "name": p.name, "team": home_name,
                "position": p.position, "minutes": p.minutes_played,
                "rating": p.rating or 0, "goals": p.goals, "assists": p.assists,
                "shots_total": p.shots_total, "shots_on": p.shots_on,
                "xg": p.xg, "xgot": p.xgot,
                "passes_total": p.passes_total, "passes_accuracy": p.passes_accuracy,
                "passes_key": p.passes_key, "passes_final_third": p.passes_final_third,
                "tackles_total": p.tackles_total, "tackles_interceptions": p.tackles_interceptions,
                "duels_won": p.duels_won, "duels_total": p.duels_total,
                "dribbles_success": p.dribbles_success, "dribbles_attempts": p.dribbles_attempts,
                "saves": p.saves, "saves_inside_box": p.saves_inside_box,
                "error_lead_to_goal": p.error_lead_to_goal,
                "photo_url": p.photo_url, "man_of_match": p.man_of_match,
                "fouls_drawn": p.fouls_drawn, "fouls_committed": p.fouls_committed,
                "ball_recoveries": p.ball_recoveries,
                "yellowcards": p.yellowcards, "redcards": p.redcards,
                "is_substitute": p.is_substitute,
            })
        for p in raw.away_players:
            if p.minutes_played <= 0:
                continue
            players_data.append({
                "name": p.name, "team": away_name,
                "position": p.position, "minutes": p.minutes_played,
                "rating": p.rating or 0, "goals": p.goals, "assists": p.assists,
                "shots_total": p.shots_total, "shots_on": p.shots_on,
                "xg": p.xg, "xgot": p.xgot,
                "passes_total": p.passes_total, "passes_accuracy": p.passes_accuracy,
                "passes_key": p.passes_key, "passes_final_third": p.passes_final_third,
                "tackles_total": p.tackles_total, "tackles_interceptions": p.tackles_interceptions,
                "duels_won": p.duels_won, "duels_total": p.duels_total,
                "dribbles_success": p.dribbles_success, "dribbles_attempts": p.dribbles_attempts,
                "saves": p.saves, "saves_inside_box": p.saves_inside_box,
                "error_lead_to_goal": p.error_lead_to_goal,
                "photo_url": p.photo_url, "man_of_match": p.man_of_match,
                "fouls_drawn": p.fouls_drawn, "fouls_committed": p.fouls_committed,
                "ball_recoveries": p.ball_recoveries,
                "yellowcards": p.yellowcards, "redcards": p.redcards,
                "is_substitute": p.is_substitute,
            })

        # Build events list
        raw_events_list = []
        for ev in raw.events:
            raw_events_list.append({
                "player_name": ev.player_name,
                "event_type": ev.event_type,
                "detail": ev.detail,
                "minute": ev.time_elapsed,
                "team_id": ev.team_id,
                "assist_name": ev.assist_name,
            })

        # Detect key events
        from src.engine.key_events import detect_key_events
        goal_events_raw = [e for e in raw_events_list if e["event_type"] in ("Goal", "goal")]
        sub_events_raw = [e for e in raw_events_list if e["event_type"] in ("subst", "substitution")]
        key_events_result = detect_key_events(
            goal_events_raw, sub_events_raw,
            raw.home_team.id, raw.away_team.id,
            total_home, total_away,
        )

        # Run player detector
        from src.engine.player_insights import run_all_detectors
        # Build lineups from raw data
        lineups_list = []
        for p in raw.home_players:
            lineups_list.append({
                "player_name": p.name, "team_id": raw.home_team.id,
                "minutes_played": p.minutes_played, "position": p.position,
                "rating": p.rating or 0, "goals": p.goals, "assists": p.assists,
                "shots_on_target": p.shots_on, "shots_total": p.shots_total,
                "passes_accuracy": p.passes_accuracy, "passes_total": p.passes_total,
                "xg": p.xg, "xgot": p.xgot,
                "dribbles_success": p.dribbles_success, "dribbles_attempts": p.dribbles_attempts,
                "duels_won": p.duels_won, "duels_total": p.duels_total,
                "tackles_total": p.tackles_total, "interceptions": p.tackles_interceptions,
                "passes_key": p.passes_key, "passes_final_third": p.passes_final_third,
                "saves": p.saves, "saves_inside_box": p.saves_inside_box,
                "fouls_committed": p.fouls_committed, "fouls_drawn": p.fouls_drawn,
                "yellowcards": p.yellowcards, "redcards": p.redcards,
            })
        for p in raw.away_players:
            lineups_list.append({
                "player_name": p.name, "team_id": raw.away_team.id,
                "minutes_played": p.minutes_played, "position": p.position,
                "rating": p.rating or 0, "goals": p.goals, "assists": p.assists,
                "shots_on_target": p.shots_on, "shots_total": p.shots_total,
                "passes_accuracy": p.passes_accuracy, "passes_total": p.passes_total,
                "xg": p.xg, "xgot": p.xgot,
                "dribbles_success": p.dribbles_success, "dribbles_attempts": p.dribbles_attempts,
                "duels_won": p.duels_won, "duels_total": p.duels_total,
                "tackles_total": p.tackles_total, "interceptions": p.tackles_interceptions,
                "passes_key": p.passes_key, "passes_final_third": p.passes_final_third,
                "saves": p.saves, "saves_inside_box": p.saves_inside_box,
                "fouls_committed": p.fouls_committed, "fouls_drawn": p.fouls_drawn,
                "yellowcards": p.yellowcards, "redcards": p.redcards,
            })

        events_parsed = []
        for ev in raw.events:
            events_parsed.append({
                "player_name": ev.player_name,
                "event_type": ev.event_type,
                "minute": ev.time_elapsed,
                "team_id": ev.team_id,
                "goal_type": ev.detail,
                "assist_name": ev.assist_name,
            })
        max_min = max(p.minutes_played for p in raw.home_players + raw.away_players if p.minutes_played > 0)
        end_min = max(90, max_min) if max_min > 0 else 90

        detector_results = run_all_detectors(
            lineups_list, raw.home_team.id, raw.away_team.id,
            total_home, total_away,
            events_parsed, end_min,
            home_name=home_name, away_name=away_name,
        )

        # Build player summaries
        player_summaries = {}
        for p in raw.home_players + raw.away_players:
            if p.minutes_played > 0:
                summary_parts = []
                if p.goals:
                    summary_parts.append(f"打入{p.goals}球")
                if p.assists:
                    summary_parts.append(f"助攻{p.assists}次")
                if p.shots_on:
                    summary_parts.append(f"射正{p.shots_on}次")
                if p.passes_key:
                    summary_parts.append(f"关键传球{p.passes_key}次")
                if p.saves and p.saves >= 3:
                    summary_parts.append(f"完成{p.saves}次扑救")
                player_summaries[p.name] = "，".join(summary_parts) if summary_parts else f"出场{p.minutes_played}分钟，评分{p.rating or '-'}"

        # Select candidates
        candidates = select_candidates(
            players_data, detector_results, key_events_result,
            raw_events_list, player_summaries, max_candidates=8,
        )
        logger.info(f"  球员候选: {len(candidates)} 人")

        # Build LLM prompt
        stage_text = f"赛事：{stage_name}" if stage_name else "赛事信息暂缺"
        feature_sys, feature_user = build_player_feature_prompt(
            match_context=match_ctx,
            stage_info=stage_text,
            tactical_summary=tactical_narrative[:1500],
            pressing_summary=pressing_narrative[:1000] if pressing_narrative else tactical_narrative[:1000],
            candidates=candidates,
            match_overview=match_overview,
        )
        logger.info("调用 LLM 生成球员特稿推荐...")
        player_features_md = llm.generate(feature_sys, feature_user, max_tokens=4800)
        logger.info(f"  球员特稿长度: {len(player_features_md)} 字符")

        # Save player features Markdown
        features_path = output_dir / "player_features.md"
        with open(features_path, "w", encoding="utf-8") as f:
            f.write(player_features_md)
    except Exception as e:
        logger.warning(f"球员特稿推荐跳过: {e}")
        import traceback
        traceback.print_exc()

    # ── 6. 保存 JSON ──
    json_path = output_dir / "tactical_analysis.json"
    json_output = dict(tactical_data) if isinstance(tactical_data, dict) else {"_data": str(tactical_data)}
    json_output["_match_overview"] = {
        "system_prompt": match_overview_sys,
        "user_prompt": match_overview_user,
        "response": match_overview,
        "response_length": len(match_overview) if match_overview else 0,
    }
    json_output["_article_naming"] = {
        "raw_response": naming_text,
        "titles": article_titles,
    }
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(json_output, f, ensure_ascii=False, indent=2, default=str)
    logger.info(f"  战术 JSON: {json_path}")

    # ── 7. 保存 Excel ──
    try:
        from run import _save_tactical_excel
        xlsx_path = output_dir / "tactical_analysis.xlsx"
        _save_tactical_excel(tactical_data, tactical_narrative,
                             home_name, away_name, str(xlsx_path))
        # Append match_overview sheet
        if match_overview and match_overview_sys:
            import openpyxl
            wb = openpyxl.load_workbook(str(xlsx_path))
            ws = wb.create_sheet("比赛过程概述")
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 80
            ws['A1'] = 'System Prompt'; ws['B1'] = match_overview_sys
            ws['A2'] = 'User Prompt'; ws['B2'] = match_overview_user
            ws['A3'] = 'LLM Response'; ws['B3'] = match_overview
            wb.save(str(xlsx_path))
        logger.info(f"  战术 Excel: {xlsx_path}")
    except Exception as e:
        logger.warning(f"战术 Excel 跳过: {e}")

    # ── 8. 组装 HTML 战术报告 ──
    _build_tactical_html(
        tactical_data, tactical_narrative, tactical_image_paths,
        home_name, away_name, score, str(output_dir),
        timeline_html, timeline_png_rel, one_liner, lineup_html, raw,
        total_home, total_away, pen_home, pen_away, has_penalties,
        pressing_narrative=pressing_narrative,
        article_titles=article_titles,
        player_features_md=player_features_md,
    )
    logger.info(f"  战术 HTML: {output_dir / 'tactical_report.html'}")

    return tactical_data, tactical_narrative


def _fallback_parse_titles(text: str) -> list[dict]:
    """Fallback 解析器：不要求分角标题，直接从文本中提取所有《标题》— 理由 行。"""
    titles = []
    for line in text.strip().split("\n"):
        line = line.strip()
        if "《" not in line or "》" not in line:
            continue
        try:
            title_part = line.split("《")[1].split("》")[0].strip()
            reason = ""
            if "—" in line:
                reason = line.split("—", 1)[1].strip()
            elif "--" in line:
                reason = line.split("--", 1)[1].strip()
            titles.append({"angle": "自由角度", "title": title_part, "reason": reason})
        except Exception:
            continue
    return titles[:10]


def _build_tactical_html(
    tactical_data: dict, tactical_narrative: str,
    tactical_image_paths: dict,
    home_name: str, away_name: str, score,
    output_dir: str, timeline_html: str, timeline_png: str,
    one_liner: str, lineup_html: str, raw,
    total_home: int, total_away: int,
    pen_home: int, pen_away: int, has_penalties: bool,
    pressing_narrative: str = "",
    article_titles: list = None,
    player_features_md: str = "",
):
    """组装独立战术分析 HTML 报告 (从 run_tactical_only.py 提取)。"""
    tac_sections = _parse_narrative_sections(tactical_narrative)

    BG = "#0f1923"; FG = "#d0d8e0"
    GREEN = "#2ecc71"; BLUE = "#3498db"; RED = "#e74c3c"; GOLD = "#f1c40f"
    CARD_BG = "#162a38"; BORDER = "#1e3a4d"

    H = []
    H.append('<!DOCTYPE html><html lang="zh-CN"><head><meta charset="utf-8">')
    H.append('<meta name="viewport" content="width=device-width, initial-scale=1.0">')
    H.append(f'<title>战术分析 — {home_name} vs {away_name}</title>')
    H.append('<style>')
    H.append(f'*{{margin:0;padding:0;box-sizing:border-box}}')
    H.append(f'body{{font-family:"Microsoft YaHei","PingFang SC",sans-serif;background:{BG};color:{FG};line-height:1.8}}')
    H.append(f'.container{{max-width:960px;margin:0 auto;padding:20px}}')
    H.append(f'img{{max-width:100%;border-radius:6px}}')
    H.append(f'h2{{color:#e0e8f0;font-size:22px;border-bottom:2px solid {GREEN};padding-bottom:8px;margin:28px 0 14px}}')
    H.append(f'.scoreboard{{text-align:center;margin:20px 0 30px}}')
    H.append(f'.scoreboard .teams{{font-size:20px;color:#fff}}')
    H.append(f'.scoreboard .score{{font-size:44px;font-weight:bold;color:{GREEN};margin:0 20px}}')
    H.append(f'.insight-box{{background:{CARD_BG};border-radius:8px;padding:14px 18px;margin:12px 0;border-left:3px solid {GREEN}}}')
    H.append(f'.insight-box.gold{{border-left-color:{GOLD}}}')
    H.append(f'.insight-box h4{{color:{GREEN};margin:0 0 6px;font-size:14px}}')
    H.append(f'.insight-box p{{font-size:14px;line-height:1.9}}')
    H.append(f'.match-bg{{font-size:13px;color:#8ab4d6;margin:10px 0;text-align:center;line-height:2}}')
    H.append(f'.match-bg .stage-name{{font-weight:bold;color:{GOLD};font-size:15px}}')
    H.append(f'.match-bg .venue-name{{color:#c0d6e4}}')
    H.append(f'.match-bg .score-detail{{font-size:12px;color:#7a9ab4;margin-top:4px}}')
    H.append(f'table{{width:100%;border-collapse:collapse;font-size:11px}}')
    H.append(f'th{{text-align:left;color:#95a5a6;font-size:10px;padding:2px 4px}}')
    H.append(f'td{{padding:3px 4px;border-bottom:1px solid {BORDER}}}')
    H.append(f'.footer{{text-align:center;color:#4a6a80;font-size:12px;margin:30px 0 10px;border-top:1px solid {BORDER};padding-top:16px}}')
    H.append('</style></head><body><div class="container">')

    # Title
    if one_liner:
        H.append(f'<div style="text-align:center;color:#fff;font-size:20px;font-weight:bold;'
                 f'line-height:1.8;padding:16px 20px;margin-bottom:8px;'
                 f'background:linear-gradient(135deg,#1a2a3a,#0f1923);border-radius:10px">{one_liner}</div>')
    else:
        H.append(f'<h1 style="text-align:center;color:#fff;font-size:24px">{home_name} vs {away_name} 战术分析报告</h1>')

    # Scoreboard
    H.append('<div class="scoreboard">')
    H.append(f'<span class="teams">{home_name}</span>')
    H.append(f'<span class="score">{total_home} - {total_away}</span>')
    H.append(f'<span class="teams">{away_name}</span>')
    if has_penalties:
        H.append(f'<p style="text-align:center;font-size:12px;color:#95a5a6;margin-top:6px">'
                 f'常规+加时 {score.home}-{score.away}  |  点球 {pen_home}-{pen_away}</p>')
    H.append('</div>')

    # Match background
    stage_name = venue_name = ""
    if raw is not None:
        si = raw.stage_info or {}; vi = raw.venue_info or {}
        stage_name = si.get("name", ""); venue_name = vi.get("name", "")
        city = vi.get("city_name", ""); capacity = vi.get("capacity", "")
    if stage_name or venue_name:
        bg_lines = []
        if stage_name:
            bg_lines.append(f'<span class="stage-name">{stage_name}</span>')
        if venue_name:
            cap_str = f" · 容量 {capacity:,}" if capacity else ""
            bg_lines.append(f'<span class="venue-name">{venue_name}{cap_str}</span>')
        score_parts = [f'常规时间 {score.home}-{score.away}']
        if score.halftime_home is not None:
            score_parts.append(f'半场 {score.halftime_home}-{score.halftime_away}')
        if (score.extratime_home or 0) or (score.extratime_away or 0):
            score_parts.append(f'加时 {score.extratime_home or 0}-{score.extratime_away or 0}')
        if has_penalties:
            score_parts.append(f'点球 {pen_home}-{pen_away}')
        bg_lines.append(f'<div class="score-detail">{"  |  ".join(score_parts)}</div>')
        H.append(f'<div class="match-bg">{"<br>".join(bg_lines)}</div>')

    # Lineup
    if lineup_html:
        H.append(lineup_html)

    # Tactical sections
    profile = tac_sections.get("战术画像", "")
    if profile:
        H.append('<div class="insight-box">')
        H.append(f'<h4>战术画像</h4><p>{profile}</p></div>')
        if tactical_image_paths.get("tactical_radar"):
            H.append(f'<p style="text-align:center;margin:16px 0"><img src="{tactical_image_paths["tactical_radar"]}" alt="战术雷达图"></p>')

    deduction = tac_sections.get("战术演绎", "")
    if deduction:
        H.append(f'<div class="insight-box" style="border-left-color:{BLUE}">')
        H.append(f'<h4 style="color:{BLUE}">战术演绎</h4><p>{deduction}</p></div>')
        if tactical_image_paths.get("tactical_possession"):
            H.append(f'<p style="text-align:center;margin:10px 0 4px;font-size:12px;color:#95a5a6">▼ 控球率逐段变化</p>')
            H.append(f'<p style="text-align:center;margin:0 0 16px"><img src="{tactical_image_paths["tactical_possession"]}" alt="控球摇摆" style="width:100%"></p>')
        if tactical_image_paths.get("tactical_shots"):
            H.append(f'<p style="text-align:center;margin:10px 0 4px;font-size:12px;color:#95a5a6">▼ 时段射门分布</p>')
            H.append(f'<p style="text-align:center;margin:0 0 16px"><img src="{tactical_image_paths["tactical_shots"]}" alt="时段射门" style="width:100%"></p>')

    verification = tac_sections.get("战术验证", "")
    if verification:
        H.append(f'<div class="insight-box" style="border-left-color:{GREEN}">')
        H.append(f'<h4 style="color:{GREEN}">战术验证</h4><p>{verification}</p></div>')

    try:
        from src.visualizer.tactical_charts import generate_tactical_html_cards
        cards = generate_tactical_html_cards(tactical_data, home_name, away_name)
        H.append(cards)
    except Exception:
        pass

    game = tac_sections.get("战术博弈", "")
    if game:
        H.append(f'<div class="insight-box" style="border-left-color:{RED}">')
        H.append(f'<h4 style="color:{RED}">战术博弈</h4><p>{game}</p></div>')
        if tactical_image_paths.get("tactical_ppda"):
            H.append(f'<p style="text-align:center;margin:10px 0 4px;font-size:12px;color:#95a5a6">▼ 全场压迫强度对比</p>')
            H.append(f'<p style="text-align:center;margin:0 0 16px"><img src="{tactical_image_paths["tactical_ppda"]}" alt="PPDA对比" style="width:100%"></p>')
        if tactical_image_paths.get("tactical_ppda_timeline"):
            H.append(f'<p style="text-align:center;margin:10px 0 4px;font-size:12px;color:#95a5a6">▼ 压迫强度随时间变化</p>')
            H.append(f'<p style="text-align:center;margin:0 0 16px"><img src="{tactical_image_paths["tactical_ppda_timeline"]}" alt="PPDA时间曲线" style="width:100%"></p>')

    # ── 双方压迫分析板块 ──
    if tactical_image_paths.get("pressing_effectiveness") or tactical_image_paths.get("pressing_efficiency"):
        H.append(f'<h2 style="color:{GREEN};font-size:22px;border-bottom:2px solid {GREEN};padding-bottom:8px;margin:28px 0 14px">双方压迫分析</h2>')

        # 压迫叙事文字
        pressing_sections = _parse_narrative_sections(pressing_narrative)
        pressing_order = ["压迫布局", "压迫回报", "压迫代价"]
        for sec_title in pressing_order:
            sec_body = pressing_sections.get(sec_title, "")
            if sec_body:
                border_c = GREEN if sec_title == "压迫布局" else (BLUE if sec_title == "压迫回报" else RED)
                H.append(f'<div class="insight-box" style="border-left-color:{border_c}">')
                H.append(f'<h4 style="color:{border_c}">{sec_title}</h4><p>{sec_body}</p></div>')

        # 压迫总结
        pressing_summary = pressing_sections.get("压迫总结", "")
        if pressing_summary:
            H.append(f'<div class="insight-box gold" style="margin-bottom:14px"><p style="font-size:14px;font-weight:bold;margin:0">{pressing_summary}</p></div>')

        if tactical_image_paths.get("pressing_effectiveness"):
            H.append(f'<p style="text-align:center;margin:10px 0 4px;font-size:12px;color:#95a5a6">▼ 压迫效果图：上图为 PPDA 压迫强度曲线与进球时刻，中图为逐段 xG 柱状，下图为射正/射偏堆叠柱</p>')
            H.append(f'<p style="text-align:center;margin:0 0 16px"><img src="{tactical_image_paths["pressing_effectiveness"]}" alt="压迫效果" style="width:100%"></p>')
        if tactical_image_paths.get("pressing_efficiency"):
            H.append(f'<p style="text-align:center;margin:10px 0 4px;font-size:12px;color:#95a5a6">▼ 压迫效率图：上图为 PPDA 压迫强度曲线与效率比，中图为抢断+拦截堆叠柱，下图为犯规次数</p>')
            H.append(f'<p style="text-align:center;margin:0 0 16px"><img src="{tactical_image_paths["pressing_efficiency"]}" alt="压迫效率" style="width:100%"></p>')

    conclusion = tac_sections.get("战术定论", "")
    if conclusion:
        H.append(f'<div class="insight-box gold"><p style="font-size:16px;font-weight:bold;margin:0">{conclusion}</p></div>')

    # ── 文章标题推荐板块 ──
    if article_titles:
        H.append(f'<h2 style="color:{GREEN};font-size:22px;border-bottom:2px solid {GREEN};padding-bottom:8px;margin:28px 0 14px">🖋️ 文章标题推荐</h2>')
        # Group by angle
        angle_colors = {
            "战术向": "#2ecc71",
            "人物向": "#e67e22",
            "数据向": "#3498db",
            "自由角度": "#95a5a6",
        }
        H.append(f'<div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:10px;margin:12px 0">')
        for t in article_titles:
            angle = t.get("angle", "自由角度")
            border_c = angle_colors.get(angle, "#95a5a6")
            H.append(f'<div style="background:rgba(46,204,113,0.06);border:1px solid {border_c};border-radius:8px;padding:12px 14px">')
            H.append(f'<div style="display:flex;align-items:center;gap:6px;margin-bottom:4px">')
            H.append(f'<span style="background:{border_c};color:#fff;font-size:10px;font-weight:bold;padding:2px 6px;border-radius:3px">{angle}</span>')
            H.append(f'</div>')
            H.append(f'<div style="font-size:15px;font-weight:700;color:#fff;line-height:1.4;margin-bottom:4px">《{t.get("title", "")}》</div>')
            reason = t.get("reason", "")
            if reason:
                H.append(f'<div style="font-size:12px;color:#8ab4d6">{reason}</div>')
            H.append(f'</div>')
        H.append(f'</div>')

    # ── 球员人物特稿推荐板块 ──
    if player_features_md:
        H.append(f'<h2 style="color:{GREEN};font-size:22px;border-bottom:2px solid {GREEN};padding-bottom:8px;margin:28px 0 14px">📰 球员人物特稿推荐</h2>')
        # Parse markdown into structured HTML
        features_html = _md_to_player_features_html(player_features_md)
        H.append(features_html)

    if timeline_html:
        H.append(f'<h2 style="color:#e0e8f0;font-size:22px;border-bottom:2px solid {GREEN};padding-bottom:8px;margin:28px 0 14px">全场事件时间轴</h2>')
        H.append(timeline_html)
        if timeline_png:
            H.append(f'<p style="text-align:center;margin:16px 0 4px;font-size:12px;color:#95a5a6">▼ 时间轴图片版（可右键保存）</p>')
            H.append(f'<p style="text-align:center;margin:0 0 16px"><img src="{timeline_png}" alt="事件时间轴" style="width:100%"></p>')

    H.append('<div class="footer">战术分析报告由 AI 自动生成 | 数据来源：SportMonks API</div>')
    H.append('</div></body></html>')

    html_path = Path(output_dir) / "tactical_report.html"
    with open(html_path, "w", encoding="utf-8") as f:
        f.write("\n".join(H))


def _parse_narrative_sections(text: str) -> dict:
    sections = {}
    for m in re.finditer(r"【(.+?)】\s*\n(.*?)(?=\n【|\Z)", text, re.DOTALL):
        sections[m.group(1).strip()] = m.group(2).strip()
    return sections


def _md_to_player_features_html(md: str) -> str:
    """Parse player features markdown into collapsible HTML cards."""
    parts = []
    # Split by ### player entries
    player_blocks = re.split(r'\n### ', md)
    overview = ""
    if player_blocks and not player_blocks[0].startswith('###'):
        overview = player_blocks[0]
        player_blocks = player_blocks[1:]

    if overview:
        # Render overview text
        for line in overview.strip().split('\n'):
            line = line.strip()
            if line.startswith('# '):
                parts.append(f'<h3 style="color:#e0e8f0;margin:10px 0">{line[2:]}</h3>')
            elif line.startswith('## '):
                parts.append(f'<h4 style="color:#c0d6e4;margin:8px 0">{line[3:]}</h4>')
            elif line:
                parts.append(f'<p style="font-size:14px;line-height:1.9;color:#d0d8e0;margin:4px 0">{line}</p>')

    for block in player_blocks:
        lines = block.strip().split('\n')
        header = lines[0].strip()
        # Parse header: "Player Name | Team | Position | ⭐..."
        header_parts = [h.strip() for h in header.split('|')]
        player_name = header_parts[0] if header_parts else ""
        extra = " | ".join(header_parts[1:]) if len(header_parts) > 1 else ""

        card_id = f"pf_{abs(hash(player_name)) % 100000}"

        parts.append(f'<div style="background:#162a38;border:1px solid #1e3a4d;border-radius:10px;margin:16px 0;overflow:hidden">')
        # Header
        parts.append(f'<div onclick="document.getElementById(\'{card_id}\').style.display='
                     f'(\'none\'==document.getElementById(\'{card_id}\').style.display)?\'\':\'none\'" '
                     f'style="cursor:pointer;padding:14px 18px;display:flex;align-items:center;gap:10px;'
                     f'background:linear-gradient(135deg,#1a2a3a,#162a38)">')
        parts.append(f'<span style="color:#e0e8f0;font-size:16px;font-weight:bold;flex:1">{player_name}</span>')
        if extra:
            parts.append(f'<span style="color:#8ab4d6;font-size:13px">{extra}</span>')
        parts.append(f'<span style="color:#4a6a80;font-size:12px">▼</span>')
        parts.append(f'</div>')

        # Collapsible detail
        parts.append(f'<div id="{card_id}" style="display:none;padding:14px 18px;border-top:1px solid #1e3a4d">')

        # Parse inner markdown
        inner_lines = lines[1:]
        in_outline = False
        outline_lines = []
        for line in inner_lines:
            line_stripped = line.strip()
            if line_stripped.startswith('**文章看点**'):
                parts.append(f'<p style="font-size:13px;color:#8ab4d6;margin:6px 0">'
                             f'<strong>文章看点</strong>：{line_stripped[len("**文章看点**"):].strip("：: ")}</p>')
            elif line_stripped.startswith('**故事线**'):
                parts.append(f'<p style="font-size:13px;color:#8ab4d6;margin:6px 0">'
                             f'<strong>故事线</strong>：{line_stripped[len("**故事线**"):].strip("：: ")}</p>')
            elif line_stripped.startswith('**文章标题备选**'):
                parts.append(f'<p style="font-size:13px;color:#f1c40f;margin:10px 0 4px"><strong>文章标题备选</strong>：</p>')
            elif line_stripped.startswith('**文章大纲**'):
                parts.append(f'<p style="font-size:13px;color:#3498db;margin:10px 0 4px"><strong>文章大纲</strong>：</p>')
                in_outline = True
            elif line_stripped.startswith('**推荐理由**'):
                parts.append(f'<p style="font-size:12px;color:#7a9ab4;margin:6px 0">'
                             f'{line_stripped}</p>')
            elif line_stripped.startswith('1. ') or line_stripped.startswith('2. ') or \
                 line_stripped.startswith('3. ') or line_stripped.startswith('4. ') or \
                 line_stripped.startswith('5. '):
                parts.append(f'<p style="font-size:12px;color:#c0d6e4;margin:2px 0 2px 12px">{line_stripped}</p>')
            elif line_stripped.startswith('- '):
                if in_outline:
                    parts.append(f'<p style="font-size:12px;color:#c0d6e4;margin:2px 0 2px 12px">{line_stripped}</p>')
                else:
                    parts.append(f'<p style="font-size:12px;color:#c0d6e4;margin:2px 0 2px 12px">{line_stripped}</p>')
            elif line_stripped == '---':
                parts.append(f'<hr style="border-color:#1e3a4d;margin:10px 0">')
            elif line_stripped and not line_stripped.startswith('#'):
                parts.append(f'<p style="font-size:13px;color:#d0d8e0;line-height:1.8;margin:4px 0">{line_stripped}</p>')

        parts.append(f'</div></div>')

    return '\n'.join(parts)


# ═══════════════════════════════════════════════════════════════
# 管道 B: 球员贡献检测器 V6
# ═══════════════════════════════════════════════════════════════

def generate_player_contribution_v6(raw, config: dict, output_dir: Path):
    """运行球员贡献检测器 V6 (JSON + Excel + 球员卡片)。"""
    logger.info("=" * 40)
    logger.info("管道 B: 球员贡献检测器 V6")

    home_name = raw.home_team.name
    away_name = raw.away_team.name
    score = raw.score
    match_id = raw.match_id

    # ── 加载 v6 模块 ──
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "src.engine.player_insights_v6",
        os.path.join(THIS_DIR, "src", "engine", "player_insights_v6.py"),
    )
    pi6 = importlib.util.module_from_spec(spec)
    sys.modules["src.engine.player_insights_v6"] = pi6
    spec.loader.exec_module(pi6)

    # ── 加载 v6 数据（需要纯 dict，直接读 JSON） ──
    json_path_raw = f"data/raw/{match_id}/raw_data.json"
    if not os.path.exists(json_path_raw):
        logger.warning(f"原始数据 JSON 不存在: {json_path_raw}")
        return []
    with open(json_path_raw, "r", encoding="utf-8") as f:
        raw_dict = json.load(f)

    match_name = f"{home_name} vs {away_name}"
    score_str = f"{score.home}:{score.away}"

    # ── LLM ──
    llm_client = None
    try:
        from src.generator.llm_client import LLMClient
        llm_client = LLMClient(config["llm"])
        logger.info(f"LLM 就绪: model={config['llm']['model']}")
    except Exception as e:
        logger.warning(f"LLM 不可用: {e}")

    # ── 运行检测器 ──
    logger.info("运行球员贡献检测 V6...")
    insights = pi6.run_v6(raw_dict, llm_client=llm_client)
    logger.info(f"  完成: {len(insights)} 名球员")

    # ── 保存 JSON ──
    os.makedirs("data/computed", exist_ok=True)
    json_path = f"data/computed/{match_id}_players_v6.json"
    _save_insights_json(insights, json_path)
    logger.info(f"  球员 JSON: {json_path}")

    # ── 保存 Excel ──
    try:
        spec2 = importlib.util.spec_from_file_location(
            "src.reporter.player_excel",
            os.path.join(THIS_DIR, "src", "reporter", "player_excel.py"),
        )
        pex = importlib.util.module_from_spec(spec2)
        sys.modules["src.reporter.player_excel"] = pex
        spec2.loader.exec_module(pex)

        xlsx_path = f"data/computed/{match_id}_players_v6.xlsx"
        pex.export_match_excel(insights, str(match_id), match_name, score_str, xlsx_path)
        logger.info(f"  球员 Excel: {xlsx_path}")
    except Exception as e:
        logger.warning(f"球员 Excel 跳过: {e}")

    # ── 摘要 ──
    roles = set(pi.role.name for pi in insights if pi.role)
    llm_count = sum(1 for pi in insights if pi.llm_summary)
    logger.info(f"  角色数: {len(roles)} | LLM 分析: {llm_count}")

    return insights


def _save_insights_json(insights: list, path: str):
    """序列化 PlayerInsightV6 列表为 JSON。"""
    data = []
    for pi in insights:
        contribs = {}
        for k, c in pi.contributions.items():
            raw_metrics_serializable = {}
            for mk, mv in c.raw_metrics.items():
                if isinstance(mv, dict):
                    raw_metrics_serializable[mk] = {kk: vv for kk, vv in mv.items()}
                else:
                    raw_metrics_serializable[mk] = {"value": mv}
            contribs[k] = {
                "zscore": c.zscore, "rank": c.rank, "percentile": c.percentile,
                "label": c.label, "raw_metrics": raw_metrics_serializable,
            }
        role = None
        if pi.role:
            role = {"name": pi.role.name, "confidence": pi.role.confidence,
                    "narrative": pi.role.narrative}
        eb = pi.event_bonus
        data.append({
            "name": pi.name, "player_id": pi.player_id, "number": pi.number,
            "pos": pi.pos, "team": pi.team, "team_name": pi.team_name,
            "minutes": pi.minutes, "is_substitute": pi.is_substitute,
            "contributions": contribs, "role": role, "llm_summary": pi.llm_summary,
            "events": eb.labels() if eb else [],
            "c6_label": eb.c6_label() if eb else "",
            "c6_score": eb.compute_score() if eb else 0,
        })
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ═══════════════════════════════════════════════════════════════
# 管道 C: 球员卡片 V6
# ═══════════════════════════════════════════════════════════════

def generate_player_cards_v6(match_id: int, output_dir: Path,
                             key_only: bool = False, player_filter: str = ""):
    """从 JSON 生成球员贡献卡片 PNG (需要 Playwright)。"""
    from generate_cards_v6 import (
        generate_all_cards, generate_key_cards, generate_player_card,
    )

    json_path = f"data/computed/{match_id}_players_v6.json"
    if not os.path.exists(json_path):
        logger.warning(f"球员 JSON 不存在: {json_path}，跳过卡片生成")
        return

    cards_dir = output_dir / "player_cards"
    cards_dir.mkdir(parents=True, exist_ok=True)

    try:
        with open(json_path, "r", encoding="utf-8") as f:
            players = json.load(f)
    except Exception as e:
        logger.warning(f"加载 JSON 失败: {e}")
        return

    logger.info("生成球员卡片 V6...")
    if player_filter:
        count = generate_player_card(players, player_filter, str(cards_dir), match_id)
        logger.info(f"  生成卡片: {player_filter} -> {count}")
    elif key_only:
        count = generate_key_cards(players, str(cards_dir), match_id)
        logger.info(f"  生成关键球员卡片: {count} 张")
    else:
        count = generate_all_cards(players, str(cards_dir), match_id)
        logger.info(f"  生成全部卡片: {count} 张")


# ═══════════════════════════════════════════════════════════════
# 主入口
# ═══════════════════════════════════════════════════════════════

def generate_full_report(
    match_id: int,
    skip_llm: bool = False,
    tactical_only: bool = False,
    cards_only: bool = False,
    key_cards_only: bool = False,
    player_filter: str = "",
):
    """统一生成比赛报告（战术 V2 + 球员 V6）。"""
    config = load_config()
    if skip_llm:
        config["llm"] = None

    raw = load_match_data(match_id, config)
    home_name = raw.home_team.name
    away_name = raw.away_team.name
    total_home, total_away, _, _, _ = compute_full_score(raw)

    safe_home = home_name.replace(" ", "_")
    safe_away = away_name.replace(" ", "_")
    output_dir = Path("output") / f"{match_id}_{safe_home}_vs_{safe_away}"
    output_dir.mkdir(parents=True, exist_ok=True)

    logger.info(f"{'='*60}")
    logger.info(f"比赛: {home_name} {total_home} - {total_away} {away_name}  (#{match_id})")
    logger.info(f"输出: {output_dir}")

    if cards_only:
        generate_player_cards_v6(match_id, output_dir, key_only=key_cards_only,
                                 player_filter=player_filter)
        return

    if not tactical_only:
        generate_player_contribution_v6(raw, config, output_dir)

    generate_tactical_report_v2(raw, config, output_dir)

    logger.info(f"\n{'='*60}")
    logger.info(f"完成! 报告目录: {output_dir}")
    logger.info(f"  tactical_report.html  — 战术分析报告 V2")
    logger.info(f"  tactical_analysis.json — 战术原始数据")
    logger.info(f"  tactical_analysis.xlsx — 战术 Excel")
    logger.info(f"  images/               — 战术图表")
    logger.info(f"  data/computed/        — 球员 V6 JSON / Excel")
    logger.info(f"  player_cards/         — 球员卡片 (需单独运行 --cards-only)")


def main():
    parser = argparse.ArgumentParser(
        description="统一比赛报告生成器 — 战术报告 V2 + 球员贡献检测器 V6",
    )
    parser.add_argument("match_id", type=int, nargs="+", help="比赛 ID")
    parser.add_argument("--no-llm", action="store_true", help="跳过 LLM 叙事")
    parser.add_argument("--tactical-only", action="store_true", help="仅生成战术报告")
    parser.add_argument("--cards-only", action="store_true", help="仅生成球员卡片")
    parser.add_argument("--key-cards", action="store_true", help="仅关键球员卡片")
    parser.add_argument("--player", type=str, default="", help="指定球员名生成卡片")
    args = parser.parse_args()

    for mid in args.match_id:
        generate_full_report(
            mid,
            skip_llm=args.no_llm,
            tactical_only=args.tactical_only,
            cards_only=args.cards_only,
            key_cards_only=args.key_cards,
            player_filter=args.player,
        )


if __name__ == "__main__":
    main()
