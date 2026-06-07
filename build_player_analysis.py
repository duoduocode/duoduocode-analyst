"""
Batch generate player analysis: cards + Excel + top-5 markers + LLM summaries.
Usage: python build_player_analysis.py [--dry-run] [match_id [match_id ...]]
"""
from __future__ import annotations

import argparse, json, os, sys, io, math
from pathlib import Path
from collections import defaultdict

import yaml
import requests
import numpy as np

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

sys.path.insert(0, '.')
from src.engine.player_insights import (
    run_all_detectors, DETECTOR_TAGS, classify_position, AllDetectorResults,
)
from src.collector.api_client import fetch_all, PLAYER_STAT_MAP
from src.generator.llm_client import LLMClient

REVERSE_MAP = {v: k for k, v in PLAYER_STAT_MAP.items()}
_POS_LABELS = {"G": "门将", "D": "后卫", "M": "中场", "F": "前锋"}

DETECTOR_ATTRS = {
    "D1": "D1_progression", "D2": "D2_pressing", "D3": "D3_gravity",
    "D4": "D4_tempo", "D5": "D5_twoway", "D6": "D6_timing",
    "D7": "D7_efficiency", "D8": "D8_role_deviation", "D9": "D9_connector",
    "D13": "D13_prowess",
}

# ── Card layout constants ──
PANEL_W = 9.6; MARGIN = 0.25; HEADER_H = 0.40; TAGS_H = 0.28
GRID_TOP_GAP = 0.10; CARD_PAD = 0.10; TITLE_H = 0.20; COL_HEAD_H = 0.18
METRIC_ROW_H = 0.16; FOOTNOTE_H = 0.14; PHOTO_R = 0.16; BADGE_R = 0.12


# ═══════════════════════════════════════════════════════════════
# Team Color Extraction
# ═══════════════════════════════════════════════════════════════

def extract_team_color(badge_url: str) -> str:
    """Extract dominant vibrant color from team badge. Returns hex string."""
    from PIL import Image
    try:
        session = requests.Session(); session.trust_env = False
        resp = session.get(badge_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)
        img = Image.open(io.BytesIO(resp.content)).convert("RGBA")
        img = img.resize((100, 100))
        arr = np.array(img).astype(float)

        # Use only non-transparent, non-white/non-black pixels
        alpha = arr[:, :, 3]
        mask = alpha > 100
        rgb = arr[:, :, :3][mask]
        if len(rgb) == 0:
            return "#1a5276"

        # Score each pixel: colorful & bright
        brightness = rgb.max(axis=1)
        saturation = (rgb.max(axis=1) - rgb.min(axis=1)) / (brightness + 1)
        # Prefer richer colors
        scores = saturation * brightness
        # Pick top 20% pixels and average
        threshold = np.percentile(scores, 80)
        selected = rgb[scores >= threshold]
        avg = selected.mean(axis=0).clip(0, 255)
        return f"#{int(avg[0]):02x}{int(avg[1]):02x}{int(avg[2]):02x}"
    except Exception:
        return "#1a5276"


# ═══════════════════════════════════════════════════════════════
# LLM Player Summary
# ═══════════════════════════════════════════════════════════════

def _norm(s: str) -> str:
    """Normalize for fuzzy matching: lowercase, strip accents."""
    import unicodedata
    s = unicodedata.normalize('NFKD', s.lower())
    return ''.join(c for c in s if not unicodedata.combining(c))


def generate_player_summaries(
    players_data: list[dict],
    llm_config: dict,
    match_context: str,
) -> dict[str, str]:
    """
    Generate 60-80 char summaries for all players via LLM.
    Uses prompts/player_summary.yaml (Jinja2 template).
    Returns {player_name: summary}.
    """
    if not players_data:
        return {}

    # Load prompt template
    prompt_path = os.path.join(os.path.dirname(__file__), "prompts", "player_summary.yaml")
    import yaml as _yaml
    with open(prompt_path, "r", encoding="utf-8") as f:
        prompt_cfg = _yaml.safe_load(f)

    system_prompt = prompt_cfg["system"]
    user_template = prompt_cfg["user"]

    # Render user prompt with Jinja2
    from jinja2 import Template
    template = Template(user_template)
    user_prompt = template.render(match_context=match_context, players=players_data)

    api_key = llm_config.get("api_key", "")
    if api_key.startswith("${"):
        env_key = api_key.strip("${}").strip()
        api_key = os.environ.get(env_key, api_key)
    api_key = os.environ.get("DEEPSEEK_API_KEY", api_key)
    base_url = llm_config.get("base_url", "https://api.deepseek.com/v1")
    model = llm_config.get("model", "deepseek-chat")

    for attempt in range(3):
        try:
            session = requests.Session()
            session.trust_env = False
            resp = session.post(
                f"{base_url}/chat/completions",
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": model,
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt},
                    ],
                    "temperature": llm_config.get("temperature", 0.7),
                    "max_tokens": llm_config.get("max_tokens", 3200),
                },
                timeout=60,
            )
            resp.raise_for_status()
            data = resp.json()
            content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
            finish_reason = data.get("choices", [{}])[0].get("finish_reason", "none")
            print(f"  LLM tokens: {data.get('usage', {}).get('total_tokens', '?')}, "
                  f"content: {len(content)} chars, finish: {finish_reason}")
            if not content:
                raise RuntimeError("LLM returned empty content")

            results = {}
            for line in content.strip().split("\n"):
                line = line.strip()
                if not line or (":" not in line and "：" not in line):
                    continue
                sep = ":" if ":" in line else "："
                name_raw = line.split(sep)[0].strip()
                comment = line.split(sep, 1)[1].strip()
                name = name_raw.lstrip("-*· 0123456789.").strip()
                if name and comment:
                    matched = None
                    for p in players_data:
                        if _norm(p["name"]) == _norm(name):
                            matched = p["name"]; break
                    if not matched:
                        for p in players_data:
                            pn_last = _norm(p["name"]).split()[-1]
                            n_last = _norm(name).split()[-1]
                            if pn_last == n_last:
                                matched = p["name"]; break
                    if not matched:
                        for p in players_data:
                            pn_norm = _norm(p["name"])
                            nm_norm = _norm(name)
                            if pn_norm in nm_norm or nm_norm in pn_norm:
                                matched = p["name"]; break
                    if matched:
                        results[matched] = comment[:90]
            return results
        except Exception as e:
            print(f"  LLM attempt {attempt + 1} failed: {e}")
            if attempt < 2:
                import time; time.sleep(2 ** attempt)
    print("  [WARN] LLM failed after 3 retries")
    return {}


# ═══════════════════════════════════════════════════════════════
# Player Card Generator (matplotlib, uses team color)
# ═══════════════════════════════════════════════════════════════

def _fetch_image(url: str) -> np.ndarray | None:
    if not url: return None
    try:
        session = requests.Session(); session.trust_env = False
        resp = session.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)
        resp.raise_for_status()
        img = plt.imread(io.BytesIO(resp.content))
        if img.ndim == 2: img = np.stack([img] * 3, axis=-1)
        if img.shape[-1] == 4: img = img[..., :3]
        return img
    except Exception: return None


def _circle_image(img: np.ndarray) -> np.ndarray:
    """Crop to circle. Handles float32(0-1) and uint8(0-255)."""
    h, w = img.shape[:2]; size = min(h, w)
    cy, cx = h // 2, w // 2
    y1, y2 = cy - size // 2, cy + size // 2
    x1, x2 = cx - size // 2, cx + size // 2
    cropped = img[y1:y2, x1:x2].copy()
    if cropped.dtype in (np.float32, np.float64):
        cropped = (cropped * 255).clip(0, 255)
    cropped = np.asarray(cropped, dtype=np.uint8)
    yy, xx = np.ogrid[:size, :size]
    dist = np.sqrt((yy - size / 2 + 0.5) ** 2 + (xx - size / 2 + 0.5) ** 2)
    mask = dist <= size / 2
    rgba = np.zeros((size, size, 4), dtype=np.uint8)
    if cropped.ndim == 3 and cropped.shape[2] >= 3:
        rgba[:, :, :3] = cropped[:, :, :3]
    else:
        rgba[:, :, :3] = cropped[:, :, :1] if cropped.ndim == 2 else cropped
    rgba[:, :, 3] = (mask * 255).astype(np.uint8)
    return rgba


def _darken(hex_color: str, factor: float = 0.7) -> str:
    """Darken a hex color by factor."""
    h = hex_color.lstrip('#')
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    r, g, b = int(r * factor), int(g * factor), int(b * factor)
    return f"#{r:02x}{g:02x}{b:02x}"


def _format_val(v) -> str:
    if isinstance(v, float):
        if abs(v) >= 100: return str(int(v))
        elif abs(v) >= 10: return f"{v:.1f}"
        elif abs(v) < 1 and v != 0: return f"{v:.3f}"
        else: return f"{v:.2f}"
    return str(v)


def build_sections(results, hname, aname, player_name, team_name) -> list[dict]:
    raw_by_detector = {}
    top3_by_detector = {}
    for dname, attr in DETECTOR_ATTRS.items():
        d = getattr(results, attr)
        raw_by_detector[dname] = d
        if isinstance(d, list):
            # D6 is a flat list, treat as single team
            top3_by_detector[dname] = {"team_results": {
                team_name: [{"name": r.name, "score": r.score} for r in d[:3]],
                hname: [{"name": r.name, "score": r.score} for r in d[:3]],
                aname: [{"name": r.name, "score": r.score} for r in d[:3]],
            }}
        else:
            top3_by_detector[dname] = {
                "team_results": {
                    _t: [{"name": r.name, "score": r.score} for r in rlist[:3]]
                    for _t, rlist in d.items()
                }
            }
    return _build_player_sections(raw_by_detector, top3_by_detector, player_name, team_name)


def _build_player_sections(raw_by_detector, top3_by_detector, player_name, team_name, max_metrics=5):
    sections = []
    for dname, attr in DETECTOR_ATTRS.items():
        info = top3_by_detector.get(dname)
        if not info: continue
        team_top3 = info.get("team_results", {}).get(team_name, [])
        if not any(r.get("name") == player_name for r in team_top3):
            continue
        raw_detector = raw_by_detector.get(dname, {})
        if isinstance(raw_detector, list):
            raw_results = raw_detector  # D6 is a flat list
        else:
            raw_results = raw_detector.get(team_name, [])
        player_result = next((r for r in raw_results if r.name == player_name), None)
        if player_result is None: continue

        metrics = []
        evidence = player_result.evidence
        if evidence:
            sorted_ev = sorted(evidence.items(),
                              key=lambda kv: abs(kv[1].get("contrib", kv[1].get("raw", 0)))
                              if isinstance(kv[1], dict) else abs(float(kv[1])) if isinstance(kv[1], (int, float)) else 0,
                              reverse=True)
            full_team_dict = raw_by_detector.get(dname, {})
            if isinstance(full_team_dict, list):
                all_list = sorted(full_team_dict, key=lambda r: -r.score)
            else:
                all_list = sorted([r for _t, _rl in full_team_dict.items() for r in _rl],
                                  key=lambda r: -r.score)
            overall_map = {p.name: i + 1 for i, p in enumerate(all_list)}
            team_map = {p.name: i + 1 for i, p in enumerate(raw_results)}
            for key, val in sorted_ev[:max_metrics]:
                raw_val = val.get("raw", val) if isinstance(val, dict) else val
                tr = team_map.get(player_name, 1)
                otr = overall_map.get(player_name, 1)
                if dname == "D5":
                    if key == "进攻z": key = "进攻贡献"
                    elif key == "防守z": key = "防守贡献"
                metrics.append((key, raw_val, tr, otr))
        sections.append({
            "tag": DETECTOR_TAGS.get(dname, dname),
            "score": round(player_result.score, 2),
            "metrics": metrics,
        })
    return sections


def render_player_card(
    player_name: str, photo_url: str, team_name: str, team_logo_url: str,
    sections: list[dict], output_path: str,
    accent_color: str = "#2ecc71",
    jersey_number: str = "", minutes: int = 0,
    llm_summary: str = "", dpi: int = 200,
):
    import matplotlib.font_manager as _fm
    # Ensure Chinese font
    _cjk = [f.name for f in _fm.fontManager.ttflist if any(
        k in f.name.lower() for k in ["simhei", "microsoft yahei", "noto sans cjk"])]
    if _cjk:
        plt.rcParams["font.sans-serif"] = [_cjk[0], "SimHei", "Microsoft YaHei"] + plt.rcParams["font.sans-serif"]
        plt.rcParams["axes.unicode_minus"] = False
    n_sections = len(sections)
    if n_sections == 0: return

    ncols = 3 if n_sections > 2 else 2
    nrows = (n_sections + ncols - 1) // ncols
    usable_w = PANEL_W - 2 * MARGIN
    gap = 0.10; card_w = (usable_w - (ncols - 1) * gap) / ncols
    max_m = max((len(s["metrics"]) for s in sections), default=0)
    card_h = TITLE_H + COL_HEAD_H + max_m * METRIC_ROW_H + 2 * CARD_PAD
    grid_h = nrows * card_h + (nrows - 1) * gap
    total_h = MARGIN + HEADER_H + TAGS_H + GRID_TOP_GAP + grid_h + MARGIN + FOOTNOTE_H

    accent = accent_color
    accent_dark = _darken(accent)
    bg_color = "#0f0f1a"

    fig, ax = plt.subplots(figsize=(PANEL_W, total_h), dpi=dpi)
    ax.set_xlim(0, PANEL_W); ax.set_ylim(0, total_h)
    ax.axis("off"); fig.patch.set_facecolor(bg_color)

    def _yt(off): return total_h - off

    photo_img = _fetch_image(photo_url); badge_img = _fetch_image(team_logo_url)
    if photo_img is not None: circle_img = _circle_image(photo_img)

    # ── HEADER ──
    header_y = _yt(MARGIN + HEADER_H / 2)
    photo_cx = MARGIN + PHOTO_R + 0.08
    if photo_img is not None:
        ax.imshow(circle_img, extent=[photo_cx - PHOTO_R, photo_cx + PHOTO_R,
                                      header_y - PHOTO_R, header_y + PHOTO_R], zorder=5)
    else:
        ax.add_patch(plt.Circle((photo_cx, header_y), PHOTO_R,
                                facecolor="#3a3a5c", edgecolor=accent, linewidth=1.5))

    # Name line: name + jersey + minutes, then summary on right
    name_x = photo_cx + PHOTO_R + 0.10
    display_name = player_name
    if jersey_number: display_name += f"  #{jersey_number}"
    if minutes > 0: display_name += f"  {minutes}'"
    ax.text(name_x, header_y, display_name, fontsize=9, color="white",
            fontweight="bold", ha="left", va="center")

    # LLM summary — right-aligned, parallel to name
    if llm_summary:
        ax.text(PANEL_W - MARGIN - 0.05, header_y, llm_summary,
                fontsize=6.5, color=accent, ha="right", va="center",
                fontstyle="italic")

    # ── TAGS ──
    tags = [s["tag"] for s in sections]
    tag_yt = _yt(MARGIN + HEADER_H + TAGS_H / 2)
    tag_x = MARGIN + 0.05; tag_h = 0.14
    for tag_text in tags:
        tag_w = max(len(tag_text) * 0.14, 0.30)
        ax.add_patch(plt.Rectangle((tag_x, tag_yt - tag_h / 2), tag_w, tag_h,
                    facecolor=accent, alpha=0.85, edgecolor=accent_dark,
                    linewidth=0.5, zorder=3))
        ax.text(tag_x + tag_w / 2, tag_yt, tag_text, fontsize=6.5, color="white",
                ha="center", va="center", fontweight="bold")
        tag_x += tag_w + 0.06

    # ── CARDS ──
    grid_top_offset = MARGIN + HEADER_H + TAGS_H + GRID_TOP_GAP
    for si, sec in enumerate(sections):
        row, col = si // ncols, si % ncols
        cx = MARGIN + col * (card_w + gap)
        ctop = grid_top_offset + row * (card_h + gap)
        cy = _yt(ctop) - card_h
        ax.add_patch(plt.Rectangle((cx, cy), card_w, card_h,
                     facecolor="#1a1a2e", edgecolor="#2a2a4a", linewidth=0.5, zorder=1))
        ty = cy + card_h - CARD_PAD - TITLE_H / 2
        ax.text(cx + CARD_PAD, ty, sec["tag"], fontsize=6.5, color=accent,
                fontweight="bold", ha="left", va="center")
        ax.text(cx + card_w - CARD_PAD, ty, f"{sec['score']:.2f}",
                fontsize=5.5, color="#8899aa", ha="right", va="center")
        sep_y = cy + card_h - CARD_PAD - TITLE_H
        ax.plot([cx + CARD_PAD, cx + card_w - CARD_PAD], [sep_y, sep_y],
                color=accent, alpha=0.2, linewidth=0.5)
        chdr_y = sep_y - COL_HEAD_H / 2
        inner_w = card_w - 2 * CARD_PAD
        cols_x = [cx + CARD_PAD,
                  cx + CARD_PAD + inner_w * 0.44,
                  cx + CARD_PAD + inner_w * 0.66,
                  cx + CARD_PAD + inner_w * 0.80]
        col_centers = [cols_x[i] + [0.44, 0.22, 0.14, 0.20][i] * inner_w / 2 for i in range(4)]
        for i, hdr in enumerate(["指标", "值", "队", "场"]):
            ax.text(col_centers[i], chdr_y, hdr, fontsize=5.0,
                    color="#556677", ha="center", va="center")
        row_bottom = sep_y - COL_HEAD_H
        for mi, (mname, mval, tr, otr) in enumerate(sec["metrics"]):
            ry = row_bottom - (mi + 0.5) * METRIC_ROW_H
            ax.text(cols_x[0] + 0.02, ry, mname, fontsize=5.5,
                    color="#ccd6e0", ha="left", va="center")
            ax.text(col_centers[1], ry, _format_val(mval), fontsize=5.5,
                    color="white", ha="center", va="center", fontweight="bold")
            ax.text(col_centers[2], ry, str(tr), fontsize=5.5, color="#aabbcc", ha="center", va="center")
            ax.text(col_centers[3], ry, str(otr), fontsize=5.5, color="#aabbcc", ha="center", va="center")

    # ── FOOTER ──
    fny = MARGIN + FOOTNOTE_H / 2
    ax.text(MARGIN + 0.05, fny, "队: 队内排名    场: 对阵双方排名",
            fontsize=4.5, color="#445566", ha="left", va="center")
    badge_cx = PANEL_W - MARGIN - BADGE_R - 0.05
    if badge_img is not None:
        ax.imshow(badge_img, extent=[badge_cx - BADGE_R, badge_cx + BADGE_R,
                                     fny - BADGE_R, fny + BADGE_R], zorder=5)
    team_x = badge_cx - BADGE_R - 0.05
    ax.text(team_x, fny, team_name, fontsize=6.5, color="#ccd6e0",
            fontweight="bold", ha="right", va="center")

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    fig.savefig(output_path, dpi=dpi, facecolor=bg_color, edgecolor="none",
                bbox_inches="tight", pad_inches=0.08)
    plt.close(fig)


# ═══════════════════════════════════════════════════════════════
# Data helpers
# ═══════════════════════════════════════════════════════════════

def player_to_lineup(p, team_id):
    details = []
    for field_name, value in p.items():
        type_id = REVERSE_MAP.get(field_name)
        if type_id and value is not None:
            details.append({'type_id': type_id, 'data': {'value': value}})
    return {
        'player_id': p['id'], 'player_name': p['name'],
        'team_id': team_id,
        'position_id': p.get('grid', p.get('position_id', 0)),
        'details': details,
        'player': {'image_path': p.get('photo_url', ''),
                   'position_id': p.get('grid', p.get('position_id', 0))},
    }


def load_or_fetch(match_id):
    raw_path = f'data/raw/{match_id}/raw_data.json'
    if os.path.exists(raw_path):
        raw = json.load(open(raw_path, 'r', encoding='utf-8'))
        # Verify it's a valid dict (not corrupted)
        if isinstance(raw, dict) and 'home_team' in raw and 'away_team' in raw:
            return raw
        # Corrupted or empty — re-fetch
        os.remove(raw_path)
    print(f"  Fetching fixture {match_id}...")
    config = yaml.safe_load(open("config.yaml", encoding="utf-8"))
    raw = fetch_all(match_id, config["sportmonks"])
    # Convert dataclass to dict for JSON serialization
    from dataclasses import asdict
    raw_dict = asdict(raw)
    os.makedirs(os.path.dirname(raw_path), exist_ok=True)
    json.dump(raw_dict, open(raw_path, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
    return raw_dict


def parse_events(RAW):
    events = []
    for e in RAW.get('events', []):
        events.append({
            'type_id': e.get('type_id', 0), 'player_id': e.get('player_id', 0),
            'participant_id': e.get('participant_id', 0),
            'related_player_id': e.get('related_player_id', 0),
            'minute': e.get('time_elapsed', 0) or e.get('minute', 0),
            'period_id': e.get('period_id', 1),
            'event_type': e.get('event_type', ''), 'detail': e.get('detail', ''),
            'player_name': e.get('player_name', ''),
            'related_player_name': e.get('related_player_name', ''),
        })
    return events


def build_player_index(RAW, home_id, away_id, hname, aname, home_logo, away_logo):
    """Build player index from raw data. Returns {name_lower: info_dict}."""
    raw_players = RAW.get('home_players', []) + RAW.get('away_players', [])
    raw_by_id = {rp['id']: rp for rp in raw_players}
    all_players = {}
    for rp in raw_players:
        pid = rp['id']; pname = rp['name']
        tid = rp.get('team_id', home_id if rp in RAW.get('home_players', []) else away_id)
        is_home = tid == home_id
        pos_code = rp.get('position', '?')
        pos = _POS_LABELS.get(pos_code, pos_code)
        all_players[pname.lower()] = {
            'name': pname, 'pid': pid,
            'photo_url': rp.get('photo_url', ''),
            'team': hname if is_home else aname,
            'team_logo': home_logo if is_home else away_logo,
            'team_color': "",  # populated later
            'position': pos, 'is_home': is_home,
            'number': str(rp.get('number', '')),
            'minutes': int(rp.get('minutes_played', 0) or 0),
            'rating': rp.get('rating') or 0,
        }
    return all_players


# ═══════════════════════════════════════════════════════════════
# Excel Export
# ═══════════════════════════════════════════════════════════════

def build_excel(results, all_players, hname, aname, home_color, away_color,
                summaries, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook(); ws = wb.active; ws.title = "球员贡献"
    headers = ["球队", "球员", "出场时间", "检测器标签", "检测器得分",
               "指标名", "指标值", "指标队排", "指标场排", "得分队排", "得分场排", "点评"]
    thin_border = Border(left=Side(style='thin', color='333333'),
                         right=Side(style='thin', color='333333'),
                         top=Side(style='thin', color='333333'),
                         bottom=Side(style='thin', color='333333'))
    hdr_font = Font(bold=True, size=10, color='ffffff')
    hdr_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center'); c.border = thin_border

    row = 2
    # Sort: top-5 count desc, then by team
    top5_players = set()
    # We'll compute top-5 after building all rows

    rows_data = []
    for dname in DETECTOR_ATTRS:
        d = getattr(results, DETECTOR_ATTRS[dname])
        # D6 is a flat list, normalize to dict for uniform processing
        if isinstance(d, list):
            d = {"": d}  # empty team key
        if not isinstance(d, dict):
            continue
        tag = DETECTOR_TAGS.get(dname, dname)
        for team, rlist in d.items():
            for r in rlist:
                info = all_players.get(r.name.lower(), {})
                if info.get('minutes', 0) <= 0:
                    continue
                actual_team = info.get('team', team)
                summary = summaries.get(r.name, "")
                score = round(r.score, 2)
                for ev_key, ev_val in r.evidence.items():
                    raw_val = ev_val.get("raw", ev_val) if isinstance(ev_val, dict) else ev_val
                    rows_data.append([actual_team, r.name,
                                     info.get('minutes', 0), tag, score,
                                     ev_key, _format_val_cell(raw_val),
                                     0, 0, 0, 0, summary])

    # ── Rankings per (tag, metric_name) by metric value ──
    from collections import defaultdict as dd
    # (tag, metric_name) -> [(team, name, raw_value), ...]
    metric_entries = dd(list)
    for rd in rows_data:
        tag, mname, team, name, mval = rd[3], rd[5], rd[0], rd[1], rd[6]
        metric_entries[(tag, mname)].append((team, name, mval))

    team_rank_map = {}    # (tag, metric, team, name) -> rank
    overall_rank_map = {} # (tag, metric, name) -> rank

    for (tag, mname), entries in metric_entries.items():
        # Overall rank: sort by value desc
        entries.sort(key=lambda x: -_num(x[2]))
        for i, (t, n, v) in enumerate(entries):
            overall_rank_map[(tag, mname, n)] = i + 1
        # Team rank
        teams = set(t for t, _, _ in entries)
        for team in teams:
            team_list = sorted([(t, n, v) for t, n, v in entries if t == team],
                               key=lambda x: -_num(x[2]))
            for i, (t, n, v) in enumerate(team_list):
                team_rank_map[(tag, mname, t, n)] = i + 1

    for rd in rows_data:
        rd[7] = team_rank_map.get((rd[3], rd[5], rd[0], rd[1]), 0)
        rd[8] = overall_rank_map.get((rd[3], rd[5], rd[1]), 0)

    # ── Rankings per detector tag by detector score ──
    # (tag) -> [(team, name, score), ...]
    score_entries = dd(list)
    score_seen = dd(set)
    for rd in rows_data:
        key = (rd[3], rd[0], rd[1])  # tag, team, name
        if key not in score_seen[rd[3]]:
            score_seen[rd[3]].add(key)
            score_entries[rd[3]].append((rd[0], rd[1], rd[4]))  # team, name, score

    score_team_map = {}     # (tag, team, name) -> rank
    score_overall_map = {}  # (tag, name) -> rank

    for tag, entries in score_entries.items():
        # D5 全能战士 uses harmonic mean where lower=better
        reverse = tag != '全能战士'
        entries.sort(key=lambda x: -x[2] if reverse else x[2])
        for i, (t, n, s) in enumerate(entries):
            score_overall_map[(tag, n)] = i + 1
        for team in set(t for t, _, _ in entries):
            team_list = sorted([(t, n, s) for t, n, s in entries if t == team],
                               key=lambda x: -x[2] if reverse else x[2])
            for i, (t, n, s) in enumerate(team_list):
                score_team_map[(tag, t, n)] = i + 1

    for rd in rows_data:
        rd[9] = score_team_map.get((rd[3], rd[0], rd[1]), 0)
        rd[10] = score_overall_map.get((rd[3], rd[1]), 0)

    # Write rows
    for rd in rows_data:
        team, pname, mins, tag, score, mname, mval, mtr, motr, str_, sotr, summary = rd
        vals = [team, pname, f"{mins}'", tag, f"{score:.2f}", mname, str(mval), mtr, motr, str_, sotr, summary]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = thin_border
            c.alignment = Alignment(horizontal='center' if ci in (1, 3, 4, 7, 8, 9, 10, 11) else 'left')
            if ci == 1:
                c.fill = PatternFill(start_color='1a5c2e' if team == hname else '1a3a6e',
                                     end_color='1a5c2e' if team == hname else '1a3a6e',
                                     fill_type='solid')
                c.font = Font(color='ffffff', bold=True, size=9)
            elif ci == 2:
                c.font = Font(bold=True, size=9)
        row += 1

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 9
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 9
    ws.column_dimensions['I'].width = 9
    ws.column_dimensions['J'].width = 9
    ws.column_dimensions['K'].width = 9
    ws.column_dimensions['L'].width = 22

    # Save with fallback if file is locked by another process
    try:
        wb.save(output_path)
    except PermissionError:
        import time as _time
        alt = output_path.replace(".xlsx", f".{int(_time.time())}.xlsx")
        wb.save(alt)
        print(f"  [NOTE] Original locked, saved to: {alt}")


def _format_val_cell(v):
    if isinstance(v, float):
        if abs(v) >= 100: return int(v)
        elif abs(v) >= 10: return round(v, 1)
        elif abs(v) < 1 and v != 0: return round(v, 3)
        else: return round(v, 2)
    return v


def _num(v) -> float:
    """Safe numeric conversion for ranking."""
    if isinstance(v, (int, float)): return float(v)
    try: return float(v)
    except (TypeError, ValueError): return 0.0


# ═══════════════════════════════════════════════════════════════
# Main pipeline
# ═══════════════════════════════════════════════════════════════

def process_match(match_id, llm_config, dry_run=False):
    print(f"\n{'='*60}\nProcessing match #{match_id}")
    RAW = load_or_fetch(match_id)
    home_id = RAW['home_team']['id']; away_id = RAW['away_team']['id']
    hname = RAW['home_team']['name']; aname = RAW['away_team']['name']
    home_logo = RAW['home_team'].get('logo_url', '')
    away_logo = RAW['away_team'].get('logo_url', '')
    score_home = RAW['score']['home']; score_away = RAW['score']['away']

    # Extract team colors
    home_color = extract_team_color(home_logo) if home_logo else "#2ecc71"
    away_color = extract_team_color(away_logo) if away_logo else "#3498db"
    print(f"  Colors: {hname}={home_color}  {aname}={away_color}")

    # Lineups
    lineups = [player_to_lineup(p, home_id) for p in RAW['home_players']]
    lineups += [player_to_lineup(p, away_id) for p in RAW['away_players']]
    events = parse_events(RAW)
    max_min = max(p.get('minutes_played', 0) or 0
                  for p in RAW.get('home_players', []) + RAW.get('away_players', []))
    end_min = max(90, max_min) if max_min > 0 else 90

    # Run detectors
    print("  Running detectors...")
    results = run_all_detectors(lineups, home_id, away_id, score_home, score_away,
                                events, end_min, home_name=hname, away_name=aname)

    # Build player index
    all_players = build_player_index(RAW, home_id, away_id, hname, aname, home_logo, away_logo)
    for k, v in all_players.items():
        v['team_color'] = home_color if v['is_home'] else away_color

    # Collect player tags (top-3 per detector)
    player_tags = defaultdict(list)
    for dname, attr in DETECTOR_ATTRS.items():
        d = getattr(results, attr)
        if isinstance(d, dict):
            for team, rlist in d.items():
                for r in rlist[:3]:
                    if r.score > 0:
                        tag = DETECTOR_TAGS.get(dname, dname)
                        player_tags[r.name].append(tag)
        elif isinstance(d, list):
            for r in d[:3]:
                if r.score > 0:
                    tag = DETECTOR_TAGS.get(dname, dname)
                    player_tags[r.name].append(tag)

    # LLM summaries
    print("  Calling LLM for player summaries...")
    match_context = f"比赛: {hname} {score_home}-{score_away} {aname}"
    players_for_llm = []
    for k, v in all_players.items():
        if v['minutes'] <= 0 or v['name'] not in player_tags:
            continue
        # Build key_stats from detector evidence (top 5 metrics across all detectors)
        all_ev = {}
        for dname, attr in DETECTOR_ATTRS.items():
            d = getattr(results, attr)
            # D6 is a flat list, normalize
            if isinstance(d, list):
                team_results = d
            else:
                team_results = d.get(v['team'], [])
            pr = next((r for r in team_results if r.name == v['name']), None)
            if pr and pr.evidence:
                for evk, evv in pr.evidence.items():
                    raw = evv.get("raw", evv) if isinstance(evv, dict) else evv
                    contrib = abs(evv.get("contrib", raw)) if isinstance(evv, dict) else abs(float(raw) if isinstance(raw, (int, float)) else 0)
                    all_ev[evk] = (raw, contrib)
        top_ev = sorted(all_ev.items(), key=lambda x: -x[1][1])[:5]
        key_stats = "，".join(f"{k} {_format_val(v[0])}" for k, v in top_ev)
        players_for_llm.append({
            'name': v['name'], 'team': v['team'],
            'position': v['position'], 'minutes': v['minutes'],
            'rating': round(v['rating'], 1) if v.get('rating') else 'N/A',
            'tags': player_tags.get(v['name'], []),
            'key_stats': key_stats or '无特殊数据',
        })
    # Sort: most tags first
    players_for_llm.sort(key=lambda x: -len(x['tags']))

    summaries = {}
    if not dry_run and players_for_llm:
        summaries = generate_player_summaries(players_for_llm, llm_config, match_context)
    print(f"  Got {len(summaries)} summaries")

    # Output dirs
    safe_dir = f"{match_id}_{hname.replace(' ', '_')}_vs_{aname.replace(' ', '_')}"
    card_dir = f"output/{safe_dir}/player_cards"
    os.makedirs(card_dir, exist_ok=True)

    # Generate cards — all players with >0 minutes
    card_count = 0
    for k, info in sorted(all_players.items(), key=lambda x: -x[1]['minutes']):
        if info['minutes'] <= 0:
            continue
        sections = build_sections(results, hname, aname, info['name'], info['team'])
        if not sections:
            continue

        summary = summaries.get(info['name'], "")
        color = info['team_color']
        fname = f"{info['name']}_card.png"
        output_path = f"{card_dir}/{fname}"

        if not dry_run:
            render_player_card(
                player_name=info['name'], photo_url=info['photo_url'],
                team_name=info['team'], team_logo_url=info['team_logo'],
                sections=sections, output_path=output_path,
                accent_color=color, jersey_number=info['number'],
                minutes=info['minutes'], llm_summary=summary,
            )
        card_count += 1
    print(f"  Generated {card_count} cards in {card_dir}")

    # Generate Excel
    xlsx_path = f"output/{safe_dir}/player_insights.xlsx"
    if not dry_run:
        build_excel(results, all_players, hname, aname, home_color, away_color,
                    summaries, xlsx_path)
    print(f"  Excel saved to {xlsx_path}")

    return card_count


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("match_ids", nargs="*", type=int,
                       help="Match IDs to process")
    parser.add_argument("--dry-run", action="store_true",
                       help="Skip LLM and file output")
    parser.add_argument("--no-llm", action="store_true",
                       help="Skip LLM call only")
    args = parser.parse_args()

    config = yaml.safe_load(open("config.yaml", encoding="utf-8"))
    llm_config = config.get("llm", {})

    match_ids = args.match_ids or [
        19683241, 18452325, 19683240, 19662566,
        19683232, 19683235, 19683238,
    ]

    for mid in match_ids:
        try:
            process_match(mid, llm_config, dry_run=args.dry_run or args.no_llm)
        except Exception as e:
            print(f"  [ERROR] Failed on {mid}: {e}")
            import traceback; traceback.print_exc()
